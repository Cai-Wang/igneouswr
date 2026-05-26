"""
whole_rock_core.py — 全岩地球化学绘图核心模块（门面 API）

功能：
- Excel 数据读取（标准/转置自动检测）
- 检测限字符串解析
- 元素别名映射
- 样品筛选
- 标准化（球粒陨石/原始地幔）
- 30 个图解绘图函数（均返回 fig, ax）

子模块（_chem.py / _ternary.py / _style.py / _normalize.py）被内部导入并 re-export。
用户只需 from whole_rock_core import *，无需关心内部结构。

使用：
  import sys; sys.path.insert(0, SKILL_DIR + '/scripts')
  from whole_rock_core import *
"""

import os, glob
from dataclasses import dataclass
import numpy as np
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from scipy import stats

# ── 导入子模块并 re-export ─────────────────────────────────
from _chem import feot_calc
from _ternary import (SQRT3_2, ternary_to_xy, ternary_corners,
                       draw_ternary_frame, draw_ternary_grid,
                       draw_ternary_ticks, label_ternary_vertices)
from _style import (times_prop, get_color, set_palette, set_style_preset, set_style,
                     scatter_samples, plot_samples_ternary, scatter_groups,
                     style_ax, save_fig, add_legend, set_out_dir,
                     generate_report_html,
                     MK_SIZE_SINGLE, MK_SIZE_TERNARY, MK_EDGE_COLOR,
                     MK_EDGE_WIDTH, MK_EDGE_WIDTH_T, ANNOTATE_OFFSET,
                     ANNOTATE_FONTSIZE, LEGEND_LOC,
                     TICK_LENGTH, TICK_LENGTH_M, TICK_WIDTH, SPINE_WIDTH, GRID_LW,
                     STYLE_PRESETS, COLOR_PALETTES, get_group_colors,
                     DEFAULT_OUT_DIR)

# ============================================================
# Excel 读取
# ============================================================
ELEM_ALIAS = {
    'TFe2O3': 'TFe2O3', 'TFe\u2082O\u2083': 'TFe2O3', 'Fe2O3': 'TFe2O3', 'Fe2O3T': 'TFe2O3',
    'FeO': 'FeO', 'FeO*': 'FeO',
    'SiO2': 'SiO2', 'SiO\u2082': 'SiO2',
    'Na2O': 'Na2O', 'Na\u2082O': 'Na2O',
    'K2O': 'K2O', 'K\u2082O': 'K2O',
    'MgO': 'MgO',
    'Al2O3': 'Al2O3', 'Al\u2082O\u2083': 'Al2O3',
    'CaO': 'CaO',
    'TiO2': 'TiO2', 'TiO\u2082': 'TiO2',
}

KNOWN_ELEMENTS = {
    'SiO2','TiO2','Al2O3','TFe2O3','FeO','MnO','MgO','CaO','Na2O','K2O',
    'P2O5','LOI','TOTAL','Li','Be','Sc','V','Cr','Mn','Co','Ni','Cu','Zn',
    'Ga','Rb','Sr','Y','Zr','Nb','Mo','Sn','Cs','Ba','La','Ce','Pr','Nd',
    'Sm','Eu','Gd','Tb','Dy','Ho','Er','Tm','Yb','Lu','Hf','Ta','W','Tl','Pb','Th','U'
}

# 检测限解析策略
DL_STRATEGY_HALF = 'half'    # <0.50 → 0.25（默认）
DL_STRATEGY_ZERO = 'zero'    # <0.50 → 0.0
DL_STRATEGY_NAN  = 'nan'     # <0.50 → NaN

def _parse_value(v, dl_strategy='half'):
    """解析单元格值：检测限字符串 → 按策略处理，数字字符串 → float，杂字符串 → NaN。"""
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return np.nan
    # 检测限格式：< + 数字（也兼容 ≤）
    if s.startswith('<') or s.startswith('≤'):
        try:
            val = float(s.lstrip('<≤').strip())
            if dl_strategy == 'half':
                return val * 0.5
            elif dl_strategy == 'zero':
                return 0.0
            else:
                return np.nan
        except ValueError:
            return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def find_excel(path=None):
    """
    查找 Excel 数据文件。
    1. 用户明确指定路径 → 直接用
    2. 当前工作目录搜索 *merged*geochem*.xlsx
    3. 找到 1 个 → 用；多个 → 报告；0 个 → 报错
    """
    if path:
        if os.path.exists(path):
            return path
        raise FileNotFoundError(f"指定的 Excel 文件不存在: {path}")

    candidates = glob.glob('*merged*geochem*.xlsx')
    if not candidates:
        candidates = glob.glob('*geochem*.xlsx')
    if len(candidates) == 1:
        print(f"[whole_rock] 自动发现 Excel: {candidates[0]}")
        return candidates[0]
    elif len(candidates) > 1:
        raise FileNotFoundError(
            f"找到多个 Excel 文件: {candidates}\n请用 EXCEL 参数指定路径。")
    else:
        raise FileNotFoundError("未找到 Excel 文件，请用 EXCEL 参数指定路径。")


class GeochemData:
    """
    地球化学数据容器。
    读取 merged Excel，自动检测格式，提供元素值查询。

    用法：
        gd = GeochemData('path/to/data.xlsx', sample_filter='24TJ02')
        sio2 = gd.get('SiO2')       # np.array，已筛选
        labels = gd.labels           # 样品名列表
        all_labels = gd.all_labels   # 全部样品名
    """

    def __init__(self, path=None, sample_filter=None, dl_strategy='half', sheet=None,
                 strict_filter=True):
        """
        Args:
            path: Excel 路径，None 则自动发现
            sample_filter: 样品筛选字符串（包含即选中），None 则全选
            dl_strategy: 检测限解析策略 'half'|'zero'|'nan'
            sheet: 工作表名，None 则自动搜索
            strict_filter: True 则空匹配报错，False 则回退到全部样品
        """
        self.path = find_excel(path)
        self.sample_filter = sample_filter
        self.dl_strategy = dl_strategy
        self._load(sheet)
        # 保存全量数据副本并设置当前 view
        self._all_elem_data = {k: v.copy() for k, v in self._elem_data.items()}
        self.idxs = list(range(len(self.all_labels)))
        self.labels = list(self.all_labels)
        self._init_groups()
        if sample_filter:
            self._filter(sample_filter, strict=strict_filter)
        self._report_missing()

    def _load(self, sheet):
        wb = openpyxl.load_workbook(self.path, data_only=True)

        # 工作表选择
        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            # 按优先级尝试
            for name in ['Geochemistry', 'geochemistry', 'Sheet1']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            else:
                ws = wb[wb.sheetnames[0]]
                print(f"[whole_rock] 使用工作表: {ws.title}")

        # 格式自检测
        r1_c2 = ws.cell(1, 2).value
        r2_c1 = ws.cell(2, 1).value

        transposed = False
        if r1_c2 is not None and str(r1_c2).strip() in KNOWN_ELEMENTS:
            transposed = True
        elif r2_c1 is not None and str(r2_c1).strip() in KNOWN_ELEMENTS:
            transposed = False
        else:
            transposed = any(
                str(ws.cell(1, c).value or '').strip() in KNOWN_ELEMENTS
                for c in range(2, min(ws.max_column + 1, 20))
            )

        if transposed:
            self._load_transposed(ws)
        else:
            self._load_standard(ws)

    def _load_standard(self, ws):
        """标准格式：Row 1=样品名，Col A=元素名

        兼容两种变体：
        - 4 行表头：Row 1 样品名、Row 2 类别、Row 3 元素符号、Row 4 单位
        - 论文附件格式：Row 1 样品名、Row 2+ 元素数据（无类别/单位行）
        自动检测元素数据开始行。
        """
        from _normalize import REE_ORDER
        KNOWN_ELEMENTS_SET = set(KNOWN_ELEMENTS)

        sample_cols = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip() != '':
                sample_cols.append(c)
            else:
                if all(ws.cell(1, c + k).value is None for k in range(1, 3)):
                    break

        self.all_labels = [ws.cell(1, c).value for c in sample_cols]

        # 自动检测数据开始行：扫描 Row 2~5，找第一个已知元素名
        data_start = None
        # 也检查是否包含 'Lithology' 行
        litho_row = None
        for r in range(2, min(ws.max_row + 1, 8)):
            v = ws.cell(r, 1).value
            if v is not None:
                s = str(v).strip()
                if s in KNOWN_ELEMENTS_SET:
                    data_start = r
                    break
                elif s == 'Lithology':
                    litho_row = r

        # 如果没找到已知元素，回退到 Row 4（旧格式兼容）
        if data_start is None:
            data_start = 4

        # 处理 Lithology 行（如果有的话）：保存到 self._lithology
        self._lithology = None
        if litho_row is not None:
            self._lithology = [ws.cell(litho_row, c).value for c in sample_cols]

        elem_row = {}
        for r in range(data_start, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None:
                elem = str(v).strip()
                if elem and elem not in ('', 'Element', 'Unit', 'Major elements', 'Trace elements'):
                    # 过滤掉计算比值行（这些是衍生值，非原始测试数据）
                    if elem in ('ΣREE', '∑REE', 'Eu/Eu*', '(La/Yb)N', '(Dy/Yb)N', '(La/Sm)N', '(Gd/Yb)N'):
                        continue
                    elem_row[elem] = r

        self._elem_data = {}
        for elem, r in elem_row.items():
            canon = ELEM_ALIAS.get(elem, elem)
            if canon not in self._elem_data:
                self._elem_data[canon] = np.array(
                    [_parse_value(ws.cell(r, c).value, self.dl_strategy) for c in sample_cols]
                )

        print(f"[whole_rock] 标准格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素"
              f"{' (含 Lithology)' if self._lithology else ''}")

    def _load_transposed(self, ws):
        """转置格式：Row 1 横铺元素名，Row 4+ 每行一个样品"""
        KNOWN_SET = set(KNOWN_ELEMENTS)

        # ---- 格式检测 ----
        # 扫描 Row 1（从 Col 2 开始，前 80 列），数出元素名列数
        elem_count = 0
        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()
            if v in KNOWN_SET:
                elem_count += 1
            # 碰到连续 3 个空单元格就停——说明元素区域结束
            if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                break

        if elem_count < 3:
            print(f"[whole_rock] ⚠️ 转置检测：Row 1 仅识别 {elem_count} 个元素列，按标准格式处理")
            self._load_standard(ws)
            return

        # ---- 构造元素→列号映射 ----
        # 从 Col 2 开始扫描：Col 1 是样品标签列（样品名 / "Sample ID" / 空），非元素
        elem_cols = {}       # 元素名 → 列号
        sample_labels = []   # 样品名列号
        in_elem_section = True

        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()

            if in_elem_section:
                if v in KNOWN_SET:
                    elem_cols[v] = c
                    continue
                elif v and v not in ('', 'Major elements (wt%)', 'Trace elements (ppm)',
                                     'Minor elements (ppm)', 'Element'):
                    # 碰到第一个非元素名 → 元素区域结束，这列是样品名
                    in_elem_section = False
                    sample_labels.append(c)
                    continue
                elif not v:
                    # 空单元格——如果连续 3+ 个空，说明 Table 区域的列结束了
                    if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                        break
                    continue
            else:
                # 样品名列
                if v:
                    sample_labels.append(c)
                elif all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                    break

        # ---- 读样本行（Row 4+），逐行提取数据 ----
        for r in range(4, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue

            # 标准物质行跳过
            if label.upper().startswith(('BCR', 'BHVO', 'AGV', 'GSP', 'G-2', 'JB-', 'JA-',
                                          'JG-', 'JP-', 'JR-', 'GSB', 'NIST', 'SRM', 'SY-',
                                          'MRG', 'PCC', 'DTS', 'W-2', 'DNC', 'BIR', 'UB-N',
                                          'ACE', 'SARM', 'IMA', 'SDC', 'DL-', 'GSR', 'GSS')):
                continue

            sample_labels.append(r)  # 行号作为样品标识

        self.all_labels = []
        self._elem_data = {}

        for r in range(4, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue
            # 标准物质跳过
            if label.upper().startswith(('BCR', 'BHVO', 'AGV', 'GSP', 'G-2', 'JB-', 'JA-',
                                          'JG-', 'JP-', 'JR-', 'GSB', 'NIST', 'SRM', 'SY-',
                                          'MRG', 'PCC', 'DTS', 'W-2', 'DNC', 'BIR', 'UB-N',
                                          'ACE', 'SARM', 'IMA', 'SDC', 'DL-', 'GSR', 'GSS')):
                continue

            self.all_labels.append(label)
            for elem, c in elem_cols.items():
                canon = ELEM_ALIAS.get(elem, elem)
                val = _parse_value(ws.cell(r, c).value, self.dl_strategy)
                if canon not in self._elem_data:
                    self._elem_data[canon] = []
                self._elem_data[canon].append(val)

        # 转 np.array
        for k in self._elem_data:
            self._elem_data[k] = np.array(self._elem_data[k])

        print(f"[whole_rock] 转置格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素")

    def _filter(self, keyword, strict=True):
        """按关键字筛选样品。

        Args:
            keyword: 筛选字符串（包含即选中）
            strict: True 则空匹配抛 ValueError；False 则回退到全部样品
        """
        idxs = [i for i, l in enumerate(self.all_labels) if keyword in str(l)]
        if not idxs:
            if strict:
                raise ValueError(
                    f"筛选关键字 '{keyword}' 未匹配任何样品 "
                    f"(共 {len(self.all_labels)} 个: {self.all_labels[:5]}...)。"
                    f"请检查关键字，或传 sample_filter=None 全选。")
            else:
                print(f"[whole_rock] ⚠️ 筛选关键字 '{keyword}' 未匹配任何样品，回退到全部样品")
                self.idxs = list(range(len(self.all_labels)))
                self.labels = list(self.all_labels)
        else:
            self.idxs = idxs
            self.labels = [self.all_labels[i] for i in idxs]
            print(f"[whole_rock] 筛选 '{keyword}'：{len(idxs)}/{len(self.all_labels)} 样品")

        # 重新构建当前 view（从全量副本索引）
        self._elem_data = {}
        for elem in self._all_elem_data:
            full_arr = self._all_elem_data[elem]
            self._elem_data[elem] = np.array([full_arr[i] for i in self.idxs])
        # 同步更新 groups
        self.groups = self.infer_groups()

    def infer_groups(self):
        """从样品名自动推断分组（字母数字前缀 + 连字符规则）。

        规则：
        1. 样品名含连字符 '-' 时，取连字符前的连续字母+数字串作为组名
           例：23TEQ02-1, 23TEQ02-2, 23TEQ02-3 → 组 '23TEQ02'
               25SMH14-1, 25SMH14-2             → 组 '25SMH14'
               YK-1, YK-2                        → 组 'YK'
               SGT-1, SGT-2                      → 组 'SGT'
        2. 不含连字符时，每个样品单独一组

        Returns:
            list[str]: 与 self.labels 等长的分组标签列表
        """
        import re
        groups = []
        for lbl in self.labels:
            lbl_str = str(lbl).strip()
            # 匹配 "字母数字前缀-数字" 模式
            m = re.match(r'^([A-Za-z0-9]+)-', lbl_str)
            if m:
                groups.append(m.group(1))
            else:
                # 无匹配：单独一组用样品名
                groups.append(lbl_str)
        return groups

    def _init_groups(self):
        """初始化分组标签，存储在 self.groups 中。"""
        self.groups = self.infer_groups()

    def _report_missing(self):
        """报告关键元素缺失情况。"""
        # 只在筛选后报告
        pass

    def get(self, elem_name):
        """获取元素数组（已筛选）。找不到返回全 NaN。"""
        canon = ELEM_ALIAS.get(elem_name, elem_name)
        if canon in self._elem_data:
            return self._elem_data[canon]
        if elem_name in self._elem_data:
            return self._elem_data[elem_name]
        print(f"[whole_rock] ⚠️ 元素 '{elem_name}' 未找到")
        return np.full(len(self.labels), np.nan)

    def check_elements(self, *elems, strict=False):
        """检查所需元素是否齐全，返回缺失列表。

        Args:
            *elems: 元素名列表
            strict: True 时缺关键元素直接返回并打印错误，False 仅 warning
        """
        missing = []
        for e in elems:
            canon = ELEM_ALIAS.get(e, e)
            if canon not in self._elem_data and e not in self._elem_data:
                missing.append(e)
        if missing:
            if strict:
                print(f"[whole_rock] ❌ 缺失关键元素: {missing}，无法出图")
            else:
                print(f"[whole_rock] ⚠️ 缺失元素: {missing}，图中对应位置将为空")
        return missing

# ── 从 _normalize 模块 re-export 标准化参考值 ─────────────────
from _normalize import (CHONDRITE, REE_ORDER,
                         PRIMITIVE_MANTLE, SPIDER_ORDER,
                         normalize)

# ── 从 diagram 模块 re-export 绘图函数 ──────────────────────
from whole_rock.diagrams._classification import plot_tas, plot_k2o_sio2, plot_afm, plot_shand, plot_winchester_floyd, plot_co_th, plot_an_ab_or, plot_qapf, plot_cabanis, plot_mullen, plot_jensen, plot_oconnor_volc, plot_ohta_arai, plot_pearce1977
from whole_rock.diagrams._source import plot_ree, plot_spider, plot_pearce_2008, plot_u_th_zr_nb, plot_pearce_1983, plot_sm_yb_la_sm, plot_sc_v, plot_ba_th_la_sm, plot_zr_y_zr, plot_gdyb_dydystar, plot_dyyb_layb, plot_nb_la_th_la
from whole_rock.diagrams._evolution import plot_harker, plot_miyashiro, plot_mgno, plot_zr_covariance
from whole_rock.diagrams._tectonic import plot_meschede, plot_wood, plot_pearce_cann, plot_4panel, plot_shervais, plot_saccani_2015, plot_harris, plot_muller_kternary
from whole_rock.diagrams._xy_diagrams import (plot_tasmiddlemostplut, plot_tasmiddlemostvolc,
    plot_coxplut, plot_coxvolc, plot_pearcenorry, plot_pearce1982,
    plot_pearcegranite, plot_pearcenbthyb, plot_pearcenbtiyb,
    plot_frost, plot_whalen1, plot_whalen2, plot_whalen3, plot_sylvester,
    plot_villaseca, plot_debonba, plot_debonpq, plot_schandl, plot_batchelor,
    plot_mullerkbinary, plot_hollocher1, plot_hollocher2,
    plot_hastie, plot_maniar, plot_agrawal, plot_verma,
    plot_larocheplut, plot_larochevolc, plot_middlemostplut,
    plot_pecetaylor, plot_layb, plot_ross)

# ============================================================
# 推荐图件调度
# ============================================================
#
# ==== 图件注册表 ====
# 所有图件的元信息集中在一张表 DIAGRAM_REGISTRY 中。
# MAFIC_DIAGRAMS / FELSIC_DIAGRAMS 从注册表派生，保持向后兼容。
# 新增图件只需：写绘图函数 + 在 DIAGRAM_REGISTRY 加一条记录。

@dataclass(frozen=True)
class DiagramSpec:
    """图件规格。"""
    fn: object          # 绘图函数（如 plot_tas）
    filename: str       # 输出文件名（如 "TAS.png"）
    desc: str           # 中文描述（如 "TAS 全碱-硅分类图"）
    needed: tuple       # 必需元素（AND），全部存在才推荐
    any_of: tuple | None    # OR 条件（如 ('FeO', 'TFe2O3')），至少一个存在即可
    rock_types: tuple   # 适用岩性：("mafic",) / ("felsic",) / ("mafic", "common")


DIAGRAM_REGISTRY = [
    # ── 📋 岩石系列 / 分类 ─────────────────────────────
    DiagramSpec(plot_tas,      "TAS.png",            "TAS 全碱-硅分类图",                         ('SiO2', 'Na2O', 'K2O'), None,                        ("mafic",)),
    DiagramSpec(plot_k2o_sio2, "K2O_SiO2_PT76.png",  "K₂O–SiO₂ 钾系列分类图",                      ('SiO2', 'K2O'),        None,                        ("mafic", "felsic")),
    DiagramSpec(plot_afm,      "AFM_IB1971.png",      "AFM 钙碱性-拉斑系列判别",                    ('Na2O', 'K2O', 'MgO'), ('FeO', 'TFe2O3'),           ("mafic",)),
    DiagramSpec(plot_shand,    "Shand_ACNK_ANK.png",  "Shand A/CNK–A/NK 铝质分类图",               ('Al2O3', 'CaO', 'Na2O', 'K2O'), None,                ("felsic",)),
    DiagramSpec(plot_winchester_floyd, "Winchester_Floyd1977_NbY_ZrTiO2.png",
                "Winchester & Floyd Zr/TiO2–Nb/Y 分类图", ('Zr', 'TiO2', 'Nb', 'Y'), None,      ("mafic", "felsic")),

    # ── 🔬 源区性质 ──────────────────────────────────
    DiagramSpec(plot_ree,      "REE_chondrite.png",   "REE 球粒陨石标准化配分图",
                ('La','Ce','Pr','Nd','Sm','Eu','Gd','Tb','Dy','Ho','Er','Tm','Yb','Lu'), None,    ("mafic", "felsic")),
    DiagramSpec(plot_spider,   "Spider_PM.png",       "原始地幔标准化蛛网图",
                ('Rb','Ba','Th','U','Nb','Ta','La','Ce','Pb','Pr','Nd','Sr','Sm','Zr','Hf','Eu','Ti','Gd','Tb','Dy','Ho','Y','Er','Tm','Yb','Lu'), None, ("mafic", "felsic")),
    DiagramSpec(plot_pearce_2008, "Pearce2008_ThYb_NbYb.png",
                "Pearce Th/Yb–Nb/Yb 源区判别图", ('Th', 'Nb', 'Yb'), None,            ("mafic",)),

    # ── 🧬 岩浆演化过程 ───────────────────────────────
    DiagramSpec(plot_harker,   "Harker_6panel.png",    "Harker 六合一协变图",                       ('SiO2','MgO','Al2O3','CaO','Na2O','TiO2'), None,      ("mafic", "felsic")),
    DiagramSpec(plot_miyashiro,"Miyashiro1974_FeOtMgO_SiO2.png", "Miyashiro FeOt/MgO–SiO₂ 构造判别", ('SiO2', 'MgO'), ('FeO', 'TFe2O3'),                    ("mafic",)),
    DiagramSpec(plot_mgno,     "MgNo_vs_SiO2.png",    "Mg# vs SiO₂ 演化图",                        ('SiO2', 'MgO'),        ('FeO', 'TFe2O3'),          ("mafic", "felsic")),
    DiagramSpec(plot_zr_covariance, "Zr_covariance.png", "Zr 协变 3×3 图",                        ('Zr',),                None,                        ("mafic", "felsic")),

    # ── 🌍 构造环境判别 ──────────────────────────────
    DiagramSpec(plot_meschede,   "Meschede1986_ternary.png",      "Meschede Nb–Zr–Y 构造判别",        ('Nb', 'Zr', 'Y'),  None,                           ("mafic",)),
    DiagramSpec(plot_wood,       "Wood1980_Hf3_Th_Ta.png",         "Wood Hf/3–Th–Ta 构造判别",          ('Hf', 'Th', 'Ta'), None,                           ("mafic",)),
    DiagramSpec(plot_pearce_cann,"PearceCann1973_TiZrY.png",      "Pearce & Cann Ti–Zr–Y 构造判别",    ('Ti', 'Zr', 'Y'),  None,                           ("mafic",)),
    DiagramSpec(plot_4panel,     "V_Ti_Sc_ThNb_BaTh_4panel.png",  "四联比值构造判别图",                 ('Ti', 'V', 'Sc', 'Nb', 'Yb', 'Th', 'La', 'Sm', 'Ba'), None, ("mafic",)),
    DiagramSpec(plot_shervais,   "Shervais1982_Ti_V.png",         "Shervais Ti-V 构造判别图",           ('Ti', 'V'),            None,                           ("mafic",)),

    # -- 新加源区/分类图 (2026-05-09) -------------------
    DiagramSpec(plot_co_th,      "Co_Th_Hastie2007.png",          "Co-Th (Hastie) 系列判别图",          ('Co', 'Th'),           None,                        ("mafic",)),
    DiagramSpec(plot_u_th_zr_nb, "UTh_ZrNb_Stern2006.png",        "U/Th-Zr/Nb (Stern) 源区判别",        ('U', 'Th', 'Zr', 'Nb'), None,                        ("mafic",)),
    DiagramSpec(plot_pearce_1983, "ThYb_TaYb_Pearce1983.png",     "Th/Yb-Ta/Yb (Pearce 1983) 构造判别", ('Th', 'Ta', 'Yb'),     None,                        ("mafic",)),
    DiagramSpec(plot_sm_yb_la_sm, "SmYb_LaSm_partial_melting.png","(Sm/Yb)PM-(La/Sm)PM 部分熔融图",    ('La', 'Sm', 'Yb'),     None,                        ("mafic",)),
    DiagramSpec(plot_saccani_2015,"NbN_ThN_Saccani2015.png",      "NbN-ThN (Saccani) 构造判别",         ('Nb', 'Th'),           None,                        ("mafic",)),
    DiagramSpec(plot_sc_v,     "Sc_V_HickeyVargas2018.png",    "Sc-V (Hickey-Vargas) 氧化条件判别",  ('Sc', 'V'),            None,                        ("mafic",)),
    DiagramSpec(plot_ba_th_la_sm, "BaTh_LaSm_PearceRobinson2010.png", "Ba/Th-La/Sm 流体vs熔体判别",('Ba', 'Th', 'La', 'Sm'), None,                      ("mafic",)),
    DiagramSpec(plot_zr_y_zr,  "ZrY_Zr_Xia2014.png",           "Zr/Y vs Zr 岛弧vs大陆弧判别",        ('Zr', 'Y'),            None,                        ("mafic",)),

    # -- 新增 (2026-05-16): 新图件 ----------------------------
    DiagramSpec(plot_gdyb_dydystar, "GdYb_DyDystar_Davidson2013.png",
                "Gd/Yb vs Dy/Dy* 稀土分馏模式 (Davidson 2013)",   ('La', 'Tb', 'Dy', 'Ho', 'Yb'),  None,  ("mafic", "felsic")),
    DiagramSpec(plot_dyyb_layb,  "DyYb_LaYb_garnet_depth.png",
                "Dy/Yb vs La/Yb 石榴石源区深度判别 (Zhang 2018)", ('La', 'Dy', 'Yb'),              None,  ("mafic", "felsic")),
    DiagramSpec(plot_an_ab_or,   "An_Ab_Or_OConnor1965.png",
                "An-Ab-Or 长石分类三元图 (O'Connor 1965)",        ('Na2O', 'K2O', 'CaO', 'Al2O3', 'SiO2'), None, ("felsic",)),
    DiagramSpec(plot_qapf,       "QAPF_Streckeisen1976.png",
                "Q-A-PF 深成岩分类三元图 (Streckeisen 1976)",     ('SiO2', 'Na2O', 'K2O', 'CaO', 'Al2O3'), None, ("felsic",)),
    DiagramSpec(plot_nb_la_th_la,"NbLa_ThLa_Cabanis1986.png",
                "Nb/La vs Th/La 构造判别 (Cabanis & Lemelle 1986)", ('Nb', 'Th', 'La'),               None,  ("mafic",)),

    # -- RockPlot SVG 三角图 (2026-05-25) ---------------------
    DiagramSpec(plot_cabanis,    "Cabanis1986_LaY_Nb_ternary.png",
                "Cabanis La/10-Y/15-Nb/8 基性岩三角图",             ('La', 'Y', 'Nb'),                None,  ("mafic",)),
    DiagramSpec(plot_mullen,     "Mullen1983_TiO2_MnO_P2O5.png",
                "Mullen TiO2-MnO-P2O5 基性岩三角图",                ('TiO2', 'MnO', 'P2O5'),         None,  ("mafic",)),
    DiagramSpec(plot_jensen,     "Jensen1976_cation_ternary.png",
                "Jensen FeOt+TiO2-Al2O3-MgO 阳离子三角图",          ('MgO', 'Al2O3'),                 ('FeO', 'TFe2O3'), ("mafic",)),
    DiagramSpec(plot_oconnor_volc,"OConnor_Volc_An_Ab_Or.png",
                "O'Connor An-Ab-Or 火山岩三角图",                   ('Na2O', 'K2O', 'CaO'),          None,  ("felsic",)),
    DiagramSpec(plot_ohta_arai,  "Ohta_Arai2007_MFW.png",
                "Ohta & Arai M-F-W 俯冲带源区三角图",               ('La', 'Nb', 'Ce', 'Zr', 'Y'),   None,  ("mafic",)),
    DiagramSpec(plot_pearce1977, "Pearce1977_FeOt_MgO_Al2O3.png",
                "Pearce FeOt-MgO-Al2O3 基性岩构造三角图",           ('MgO', 'Al2O3'),                 ('FeO', 'TFe2O3'), ("mafic",)),
    DiagramSpec(plot_harris,     "Harris1986_Rb30_Hf_3Ta.png",
                "Harris Rb/30-Hf-3Ta 花岗岩构造判别三角图",         ('Rb', 'Hf', 'Ta'),               None,  ("felsic",)),
    DiagramSpec(plot_muller_kternary, "Muller2000_Kternary.png",
                "Muller Th-Ta-Hf 三子图等边三元构造判别图",         ('Th', 'Ta', 'Hf'),              None,  ("mafic",)),

    # -- XY 二元图 (2026-05-26) ---------------------------------
    DiagramSpec(plot_tasmiddlemostplut, "TAS_Middlemost1994_Plutonic.png",
                "TAS 深成岩全碱-硅分类 (Middlemost 1994)",         ('SiO2', 'Na2O', 'K2O'),           None,  ("felsic",)),
    DiagramSpec(plot_tasmiddlemostvolc, "TAS_Middlemost1994_Volcanic.png",
                "TAS 火山岩全碱-硅分类 (Middlemost 1994)",         ('SiO2', 'Na2O', 'K2O'),           None,  ("mafic",)),
    DiagramSpec(plot_coxplut,    "TAS_Cox1979_Plutonic.png",
                "TAS 深成岩分类 (Cox 1979)",                      ('SiO2', 'Na2O', 'K2O'),           None,  ("felsic",)),
    DiagramSpec(plot_coxvolc,    "TAS_Cox1979_Volcanic.png",
                "TAS 火山岩分类 (Cox 1979)",                      ('SiO2', 'Na2O', 'K2O'),           None,  ("mafic",)),
    DiagramSpec(plot_pearcenorry,"Pearce_Norry1979_ZrY_Zr.png",
                "Pearce & Norry Zr/Y–Zr WPB/MORB/IAB 构造判别",               ('Zr', 'Y'),                       None,  ("mafic",)),
    DiagramSpec(plot_pearce1982, "Pearce1982_ZrY_Zr.png",
                "Pearce (1982) Zr/Y–Zr + Ti/Nb/Sr 多元素判别",                  ('Zr', 'Y'),                       None,  ("mafic",)),
    DiagramSpec(plot_pearcegranite, "Pearce1984_Granite_Rb_YNb.png",
                "Pearce Rb–Y+Nb 花岗岩构造判别",                  ('Rb', 'Y', 'Nb'),                 None,  ("felsic",)),
    DiagramSpec(plot_pearcenbthyb, "Pearce1995_NbYb_ThYb.png",
                "Pearce Nb/Yb–Th/Yb 源区判别",                   ('Nb', 'Th', 'Yb'),                None,  ("mafic",)),
    DiagramSpec(plot_pearcenbtiyb, "Pearce1995_TiYb_NbYb.png",
                "Pearce Ti/Yb–Nb/Yb 源区判别",                   ('Ti', 'Nb', 'Yb'),                None,  ("mafic",)),
    DiagramSpec(plot_frost,      "Frost2001_Fenum_SiO2.png",
                "Frost Fe-number vs SiO₂ 铁质-镁质分类",          ('SiO2', 'MgO'),                   ('FeO', 'TFe2O3'), ("felsic",)),
    DiagramSpec(plot_whalen1,    "Whalen1987_GaAl_Zr.png",
                "Whalen 10000×Ga/Al–Zr A型花岗岩判别",           ('Ga', 'Al2O3', 'Zr'),             None,  ("felsic",)),
    DiagramSpec(plot_whalen2,    "Whalen1987_GaAl_Nb.png",
                "Whalen 10000×Ga/Al–Nb A型花岗岩判别",           ('Ga', 'Al2O3', 'Nb'),             None,  ("felsic",)),
    DiagramSpec(plot_whalen3,    "Whalen1987_GaAl_CeYZr.png",
                "Whalen 10000×Ga/Al–Ce+Y+Zr A型花岗岩判别",      ('Ga', 'Al2O3', 'Ce', 'Y', 'Zr'),  None,  ("felsic",)),
    DiagramSpec(plot_sylvester,  "Sylvester1989_CaONa2O_Al2O3.png",
                "Sylvester CaO/Na₂O–Al₂O₃ 花岗岩源区判别",       ('CaO', 'Na2O', 'Al2O3'),           None,  ("felsic",)),
    DiagramSpec(plot_villaseca,  "Villaseca1998_ASI_FMM.png",
                "Villaseca ASI–FMM 花岗岩源区分类",              ('Al2O3', 'CaO', 'Na2O', 'K2O', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_debonba,    "Debon1983_BA_diagram.png",
                "Debon B-A 花岗岩矿物分类图",                    ('Al2O3', 'K2O', 'Na2O', 'CaO', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_debonpq,    "Debon1983_PQ_diagram.png",
                "Debon P-Q 花岗岩矿物分类图",                    ('SiO2', 'Al2O3', 'K2O', 'Na2O', 'CaO', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_schandl,    "Schandl2004_Y_Zr.png",
                "Schandl Y–Zr 花岗岩构造判别图",                 ('Y', 'Zr'),                        None,  ("felsic",)),
    DiagramSpec(plot_batchelor,  "Batchelor1985_R1_R2.png",
                "Batchelor & Bowden R1-R2 花岗岩构造判别",       ('SiO2', 'Al2O3', 'K2O', 'Na2O', 'CaO', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_mullerkbinary, "Muller1992_K2O_SiO2.png",
                "Muller K₂O–SiO₂ 岩浆系列判别",                 ('SiO2', 'K2O'),                    None,  ("mafic",)),
    DiagramSpec(plot_hollocher1, "Hollocher2012_VSc.png",
                "Hollocher V/Sc–V+Sc 弧岩浆氧化条件判别",        ('V', 'Sc'),                        None,  ("mafic",)),
    DiagramSpec(plot_hollocher2, "Hollocher2012_VSc_ZrCe.png",
                "Hollocher Zr/Ce–V/Sc 弧岩浆分类",              ('V', 'Sc', 'Zr', 'Ce'),            None,  ("mafic",)),
    DiagramSpec(plot_hastie,     "Hastie2007_Th_Co.png",
                "Hastie Th–Co 弧岩浆系列分类",                   ('Th', 'Co'),                       None,  ("mafic",)),
    DiagramSpec(plot_maniar,     "Maniar1989_Granite_disc.png",
                "Maniar & Piccoli 花岗岩构造判别",               ('SiO2', 'Al2O3', 'MgO', 'CaO', 'Na2O', 'K2O', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_agrawal,    "Agrawal2004_DF1_DF2.png",
                "Agrawal DF1-DF2 基性岩构造判别",               ('TiO2', 'Al2O3', 'MgO', 'CaO', 'Na2O', 'K2O', 'MnO', 'P2O5', 'SiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("mafic",)),
    DiagramSpec(plot_verma,      "Verma_discriminant_DF1_DF2.png",
                "Verma 判别函数 基性岩构造判别",                 ('TiO2', 'Al2O3', 'MgO', 'CaO', 'Na2O', 'K2O', 'MnO', 'P2O5', 'SiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("mafic",)),
    DiagramSpec(plot_larocheplut,"LaRoche1980_R1_R2_plutonic.png",
                "La Roche R1-R2 侵入岩分类图",                  ('SiO2', 'Al2O3', 'K2O', 'Na2O', 'CaO', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("felsic",)),
    DiagramSpec(plot_larochevolc,"LaRoche1980_R1_R2_volcanic.png",
                "La Roche R1-R2 火山岩分类图",                  ('SiO2', 'Al2O3', 'K2O', 'Na2O', 'CaO', 'MgO', 'TiO2'),
                                                                  ('FeO', 'TFe2O3'),                       ("mafic",)),
    DiagramSpec(plot_middlemostplut, "Middlemost1991_Plutonic.png",
                "Middlemost Na₂O+K₂O–SiO₂ 深成岩分类",           ('SiO2', 'Na2O', 'K2O'),            None,  ("felsic",)),
    DiagramSpec(plot_pecetaylor,  "Peccerillo_Taylor1976_K2O_SiO2.png",
                "Peccerillo & Taylor K₂O–SiO₂ 岩浆系列判别",     ('SiO2', 'K2O'),                    None,  ("mafic",)),
    DiagramSpec(plot_layb,       "LaYb_vs_Yb.png",
                "La/Yb vs Yb 源区部分熔融判别",                 ('La', 'Yb'),                        None,  ("mafic", "felsic")),
    DiagramSpec(plot_ross,       "Ross2009_LaSm_LaYb.png",
                "Ross & Bédard La/Sm–La/Yb 岩浆过程判别",        ('La', 'Sm', 'Yb'),                  None,  ("mafic", "felsic")),
]

# 从注册表派生 MAFIC_DIAGRAMS / FELSIC_DIAGRAMS（保持向后兼容）
# 每个 spec → (fn, desc, needed, any_of) 四元组
def _is_mafic(d: DiagramSpec) -> bool:
    return "mafic" in d.rock_types

def _is_felsic(d: DiagramSpec) -> bool:
    return "felsic" in d.rock_types

MAFIC_DIAGRAMS = [(d.fn, d.desc, d.needed, d.any_of) for d in DIAGRAM_REGISTRY if _is_mafic(d)]
FELSIC_DIAGRAMS = [(d.fn, d.desc, d.needed, d.any_of) for d in DIAGRAM_REGISTRY if _is_felsic(d)]

# 从注册表派生文件名映射（plot_recommended 中不再需要硬编码 name_map）
_FILENAME_MAP = {d.fn.__name__: d.filename for d in DIAGRAM_REGISTRY}






def recommended_diagrams(gd, rock_type='auto'):
    """
    根据岩石类型推荐图件列表。

    Args:
        gd: GeochemData 实例
        rock_type: 'mafic', 'felsic', 或 'auto'（自动从 SiO₂ 判断）

    Returns:
        (list of (plot_fn, description), list of (desc, reason)) — 推荐图和跳过的图
    """
    if rock_type == 'auto':
        sio2 = gd.get('SiO2')
        sio2_min = np.nanmin(sio2) if not np.all(np.isnan(sio2)) else 0
        rock_type = 'mafic' if sio2_min < 52 else 'felsic'
        type_name = '镁铁质 (Mafic, SiO₂ min < 52%)' if sio2_min < 52 else '长英质 (Felsic, SiO₂ min ≥ 52%)'
    elif rock_type == 'mafic':
        type_name = '镁铁质 (Mafic)'
    else:
        type_name = '长英质 (Felsic)'

    pool = MAFIC_DIAGRAMS if rock_type == 'mafic' else FELSIC_DIAGRAMS

    results = []
    skipped = []
    for fn, desc, needed, any_of in pool:
        missing = [e for e in needed if e not in gd._elem_data]
        if missing:
            skipped.append((desc, missing))
            continue
        # any_of OR 条件：至少有一个元素存在即可
        if any_of:
            any_present = [e for e in any_of if e in gd._elem_data]
            if not any_present:
                skipped.append((desc, f"any_of {any_of} 全部缺失"))
                continue
        results.append((fn, desc))

    print(f"[whole_rock] 岩性判定: {type_name}")
    print(f"[whole_rock] 推荐图件: {len(results)} 张")
    for fn, desc in results:
        print(f"   ✓ {fn.__name__:28s} — {desc}")
    if skipped:
        print(f"[whole_rock] 因缺元素跳过: {len(skipped)} 张")
        for desc, missing in skipped:
            print(f"   ✗ 缺 {missing}  → 跳过 {desc}")

    return results, skipped


def plot_recommended(gd, out_dir=None, rock_type='auto'):
    """
    根据岩石类型一键出所有推荐图。
    内部调用 recommended_diagrams 判定岩性并过滤缺元素图，然后依次出图。

    Args:
        gd: GeochemData 实例
        out_dir: 输出目录，None 则用默认
        rock_type: 'mafic', 'felsic', 或 'auto'（自动从 SiO₂ 判断）

    Returns:
        dict — {'success': [(fn_name, file), ...], 'skipped': [(desc, missing), ...]}
    """
    from _style import DEFAULT_OUT_DIR, _OUT_DIR
    final_dir = out_dir or _OUT_DIR or DEFAULT_OUT_DIR

    # 复用 recommended_diagrams 做岩性判定 + 缺元素过滤
    diagrams, skipped_pre = recommended_diagrams(gd, rock_type=rock_type)
    success = []
    # skipped 从 recommended_diagrams 透传，后续追加运行时跳过
    skipped = list(skipped_pre)

    # diagrams 已是 recommended_diagrams 过滤后的结果，不再需要缺元素检查
    for fn, desc in diagrams:
        try:
            fig_result = fn(gd, out_dir=final_dir)
            # 一些图缺元素时返回 (None, None) 而非 raise
            if isinstance(fig_result, tuple) and len(fig_result) == 2 and fig_result[0] is None:
                skipped.append((desc, "缺关键元素（strict check 未通过）"))
                continue
            # 推断文件名：从注册表派生
            fname = _FILENAME_MAP.get(fn.__name__, f'{fn.__name__}.png')
            success.append((fn.__name__, fname))
        except Exception as e:
            skipped.append((desc, str(e)))

    print(f"\n[whole_rock] ==== 推荐出图完成 ====")
    print(f"   ✓ 成功: {len(success)} 张")
    for name, fname in success:
        print(f"      {name:28s} → {fname}")
    if skipped:
        print(f"   ✗ 跳过: {len(skipped)} 项")
        for desc, reason in skipped:
            print(f"      {desc} ({reason})")

    # 自动生成 HTML 报告
    gd_path = gd.path if hasattr(gd, 'path') else None
    try:
        generate_report_html(
            success, skipped, gd=gd, out_dir=final_dir, rock_type=rock_type)
    except Exception as e:
        print(f"[whole_rock] ⚠️ 报告生成跳过: {e}")

    return {'success': success, 'skipped': skipped}
