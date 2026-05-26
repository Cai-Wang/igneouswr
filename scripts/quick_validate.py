"""
quick_validate.py — 地球化学绘图 skill 轻量验证

覆盖核心功能的快速验证（数据读取 / 检测限解析 / 样品筛选 / 图烟雾测试）。

用法：
  python quick_validate.py [--skill-dir DIR] [--quick]

选项：
  --quick    仅做 import / registry 检查，跳过真实出图（秒级回归）

环境预检：Python >= 3.10, numpy, matplotlib, scipy, openpyxl
"""

import sys, os, tempfile

# Windows GBK 控制台下也能正确显示 UTF-8 字符
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import numpy as np

# ── Bootstrap: 找到 whole_rock_core ──────────────────────────
def _find_skill_dir():
    """向上搜索 SKILL.md 所在目录。"""
    cur = os.path.dirname(os.path.abspath(__file__))
    for _ in range(5):
        if os.path.isfile(os.path.join(cur, 'SKILL.md')):
            return cur
        cur = os.path.dirname(cur)
    return os.path.dirname(os.path.abspath(__file__))

SKILL_DIR = None  # will be set by --skill-dir or auto-detection

def _bootstrap():
    global SKILL_DIR
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--skill-dir', default=None)
    parser.add_argument('--quick', action='store_true', default=False,
                        help='仅做 import/registry 检查，跳过出图')
    args, _ = parser.parse_known_args()
    SKILL_DIR = args.skill_dir or _find_skill_dir()
    scripts_dir = os.path.join(SKILL_DIR, 'scripts')
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    return SKILL_DIR, args.quick

QUICK_MODE = _bootstrap()[1]

# ── 导入待测模块 ──────────────────────────────────────────
from whole_rock_core import (
    GeochemData, normalize,
    plot_tas, plot_k2o_sio2, plot_ree,
    recommended_diagrams, plot_recommended,
    MAFIC_DIAGRAMS, FELSIC_DIAGRAMS,
)
import _style  # noqa: E402 — for set_out_dir/set_style access via module

# ── 辅助函数 ──────────────────────────────────────────────
_PASS = 0
_FAIL = 0
_SKIP = 0

def test(name, cond, detail=""):
    global _PASS, _FAIL
    if cond:
        _PASS += 1
        print(f"  ✅ {name}")
    else:
        _FAIL += 1
        print(f"  ❌ {name}{' — ' + detail if detail else ''}")

def skip(name, reason=""):
    global _SKIP
    _SKIP += 1
    print(f"  ⏭️  {name}{' (' + reason + ')' if reason else ''}")

def _make_test_excel(elements, samples, data, sheet_name="Geochemistry", transpose=False):
    """创建测试用 Excel 文件。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if transpose:
        # 真实转置格式：元素横铺在 Row 1（Col 2+），样品在列（Row 5+）
        # Row 1: 空或 Sample ID | SiO2 | TiO2 | Al2O3 | ...
        # Row 2: 空 | wt% | wt% | ...
        # Row 3: Element | ...
        # Row 4: Unit | ...
        # Row 5+: 样品名 | 数据 | ...
        ws.cell(row=1, column=1, value="Sample ID")
        for j, elem in enumerate(elements):
            ws.cell(row=1, column=2+j, value=elem)
        # Row 2: 类别/单位占位
        for j in range(len(elements)):
            ws.cell(row=2, column=2+j, value="wt%")
        # Row 3-4: 空占位（模拟真实实验室报告）
        ws.cell(row=3, column=1, value="Element")
        ws.cell(row=4, column=1, value="Unit")
        # Data: 每行一个样品
        for i, sid in enumerate(samples):
            ws.cell(row=5+i, column=1, value=sid)
            for j, elem in enumerate(elements):
                ws.cell(row=5+i, column=2+j, value=data[elem][i])
    else:
        # 标准格式：Row 1=样品ID, Row 2=Element标题, Row 3=Unit, Row 4+=元素名称(A列)+数据
        ws.cell(row=1, column=1, value="Sample ID")
        for j, sid in enumerate(samples):
            ws.cell(row=1, column=2+j, value=sid)
        ws.cell(row=2, column=1, value="Element")
        ws.cell(row=3, column=1, value="Unit")
        for j in range(len(samples)):
            ws.cell(row=3, column=2+j, value="wt%")
        for i, elem in enumerate(elements):
            ws.cell(row=4+i, column=1, value=elem)
            for j, sid in enumerate(samples):
                ws.cell(row=4+i, column=2+j, value=data[elem][j])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    wb.close()
    return tmp.name


def _make_mini_excel():
    """创建一份最小可测试的 Excel。"""
    elements = ['SiO2', 'TiO2', 'Al2O3', 'FeO', 'TFe2O3', 'MgO', 'CaO', 'Na2O', 'K2O',
                'La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu',
                'Rb', 'Ba', 'Th', 'U', 'Nb', 'Ta', 'Pb', 'Sr', 'Zr', 'Hf', 'Y',
                'V', 'Sc', 'Ni', 'Cr', 'Co', 'Cu', 'Zn', 'Ga']
    samples = ['S01', 'S02', 'S03', 'S04', 'S05']
    np.random.seed(42)
    data = {}
    for e in elements:
        data[e] = [round(np.random.uniform(0.01, 70), 2) for _ in samples]
    # 确保 SiO2 覆盖足够范围
    data['SiO2'] = [48.5, 52.1, 55.3, 60.0, 65.2]
    data['Na2O'] = [2.1, 2.8, 3.2, 3.8, 4.1]
    data['K2O']  = [0.5, 1.2, 1.8, 2.5, 3.8]
    data['CaO']  = [8.5, 7.2, 5.8, 4.2, 2.5]
    data['MgO']  = [7.2, 5.1, 3.8, 2.1, 1.2]
    # 给部分元素加检测限
    data['La'][2] = '<0.05'
    data['Ce'][1] = '<0.10'
    return elements, samples, data


def _make_dl_excel():
    """创建含检测限字符串的 Excel。"""
    elements = ['SiO2', 'Al2O3', 'MgO', 'La', 'Ce']
    samples = ['A', 'B', 'C']
    data = {
        'SiO2':  [40.0, 45.0, 50.0],
        'Al2O3': [12.0, 15.0, 18.0],
        'MgO':   [5.0, '<0.50', 8.0],
        'La':    ['<0.01', 15.0, 22.0],
        'Ce':    [30.0, 40.0, '<0.05'],
    }
    return elements, samples, data


# ── 测试套件 ──────────────────────────────────────────────

def test_data_loading():
    print("\n[1] 数据加载 — 标准格式")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    try:
        gd = GeochemData(path)
        test("读取成功", gd is not None)
        test("样品数正确", len(gd.labels) == 5, f"got {len(gd.labels)}")
        test("SiO2 可访问", gd.get('SiO2') is not None)
        test("LA 元素存在", 'La' in gd._elem_data)
    except Exception as e:
        test("读取成功", False, str(e))
    os.unlink(path)


def test_transpose_loading():
    print("\n[2] 数据加载 — 转置格式")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data, transpose=True)
    try:
        gd = GeochemData(path)
        test("读取成功", gd is not None)
        test("样品数正确", len(gd.labels) == 5, f"got {len(gd.labels)}")
        test("SiO2 可访问", gd.get('SiO2') is not None)
    except Exception as e:
        test("读取成功", False, str(e))
    os.unlink(path)


def test_detection_limit():
    print("\n[3] 检测限解析")
    elements, samples, data = _make_dl_excel()
    path = _make_test_excel(elements, samples, data)

    # 默认策略（取半值）
    gd = GeochemData(path)
    mg = gd.get('MgO')
    test("检测限 MgO 取半值", abs(mg[1] - 0.25) < 0.01, f"got {mg[1]}")

    # zero 策略
    gd0 = GeochemData(path, dl_strategy='zero')
    mg0 = gd0.get('MgO')
    test("检测限 MgO 取零值", mg0[1] == 0.0, f"got {mg0[1]}")

    # nan 策略
    gd_nan = GeochemData(path, dl_strategy='nan')
    mg_nan = gd_nan.get('MgO')
    test("检测限 MgO 取 NaN", np.isnan(mg_nan[1]) if hasattr(mg_nan, '__len__') else True)

    os.unlink(path)


def test_sample_filter():
    print("\n[4] 样品筛选")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    gd = GeochemData(path, sample_filter='S01')
    test("筛选后样品数为 1", len(gd.labels) == 1, f"got {len(gd.labels)}")
    test("筛选后标签为 S01", gd.labels[0] == 'S01')

    # 空筛选 = 全选
    gd2 = GeochemData(path, sample_filter=None)
    test("空筛选不截断", len(gd2.labels) == 5)

    # all_labels 不受筛选影响
    test("all_labels 保留全量", len(gd.all_labels) == 5, f"got {len(gd.all_labels)}")

    os.unlink(path)


def test_normalize():
    print("\n[5] 标准化函数")
    ref = {'La': 10, 'Ce': 20}
    data_dict = {'La': 30, 'Ce': 40, 'X': 5}
    result = normalize(data_dict, ref)
    test("La 标准化正确", abs(result['La'] - 3.0) < 0.01)
    test("Ce 标准化正确", abs(result['Ce'] - 2.0) < 0.01)
    test("无参考值的元素为 NaN", np.isnan(result['X']))


def test_set_out_dir():
    print("\n[6] 输出目录设置")
    tmp = tempfile.mkdtemp()
    _style.set_out_dir(tmp)
    test("set_out_dir 设置成功", _style._OUT_DIR == tmp, f"got {_style._OUT_DIR}")


def test_set_style():
    print("\n[7] 风格定制")
    import _style
    original = _style.MK_SIZE_SINGLE
    _style.set_style(MK_SIZE_SINGLE=99)
    test("set_style 影响 _style.MK_SIZE_SINGLE", _style.MK_SIZE_SINGLE == 99, f"got {_style.MK_SIZE_SINGLE}")
    # 重置
    _style.set_style(MK_SIZE_SINGLE=original)


def test_plot_smoke():
    print("\n[8] 图件烟雾测试")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    gd = GeochemData(path)

    out = tempfile.mkdtemp()

    # K2O-SiO2
    try:
        fig, ax = plot_k2o_sio2(gd, out_dir=out)
        test("K2O-SiO2 出图", fig is not None)
        import matplotlib.pyplot as plt
        plt.close(fig)
    except Exception as e:
        test("K2O-SiO2 出图", False, str(e))

    # REE（缺关键元素时返回 None, None）
    try:
        fig2, ax2 = plot_ree(gd, out_dir=out)
        test("REE 出图", fig2 is not None)
        if fig2:
            plt.close(fig2)
    except Exception as e:
        test("REE 出图", False, str(e))

    # TAS（不依赖 pyrolite，纯 matplotlib 实现）
    try:
        fig3, ax3 = plot_tas(gd, out_dir=out)
        test("TAS 出图", fig3 is not None)
        if fig3:
            plt.close(fig3)
    except Exception as e:
        test("TAS 出图", False, str(e))

    os.unlink(path)


def test_recommended_diagrams():
    print("\n[9] 推荐列表调度")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    gd = GeochemData(path)

    # 默认 auto
    results, _ = recommended_diagrams(gd)
    test("推荐列表非空", len(results) > 0, f"got {len(results)}")

    # 手动指定 mafic
    results_m, _ = recommended_diagrams(gd, rock_type='mafic')
    test("mafic 推荐列表", len(results_m) > 0)

    # 手动指定 felsic
    results_f, _ = recommended_diagrams(gd, rock_type='felsic')
    test("felsic 推荐列表", len(results_f) > 0)

    os.unlink(path)


def test_plot_recommended():
    print("\n[10] 一键推荐出图")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    gd = GeochemData(path)
    out = tempfile.mkdtemp()

    result = plot_recommended(gd, out_dir=out)
    test("plot_recommended 返回 dict", isinstance(result, dict))
    test("结果含 success 键", 'success' in result)
    test("结果含 skipped 键", 'skipped' in result)
    # skipped 内容断言：珠网图/Ti/PearceCann 因缺 Ti 被跳过
    skipped_descs = [s[0] for s in result['skipped']]
    test("skipped 含蛛网图（缺 Ti）", any('蛛网' in d or 'Spider' in d for d in skipped_descs),
         f"skipped descs: {skipped_descs}")
    test("skipped 含 PearceCann（缺 Ti）", any('Pearce' in d for d in skipped_descs),
         f"skipped descs: {skipped_descs}")
    print(f"      success={len(result['success'])}, skipped={len(result['skipped'])}")
    # 确认 success+skipped 总数正确
    test("推荐图+跳过图 = 推荐总数",
         len(result['success']) + len(result['skipped']) == len(MAFIC_DIAGRAMS),
         f"success={len(result['success'])} + skipped={len(result['skipped'])} != {len(MAFIC_DIAGRAMS)}")

    os.unlink(path)


def test_any_of_or_condition():
    print("\n[11] any_of OR 条件 — FeO/TFe2O3 二选一")
    elements, samples, data = _make_mini_excel()
    path = _make_test_excel(elements, samples, data)
    gd = GeochemData(path)

    # 确认 AFM/Miyashiro/Mg# 被推荐（它们依赖 FeO/TFe2O3 的 any_of）
    diags, _ = recommended_diagrams(gd, rock_type='mafic')
    fn_names = {fn.__name__ for fn, _ in diags}
    test("AFM 在推荐列表中 (有 FeO)", 'plot_afm' in fn_names, f"got {fn_names}")
    test("Miyashiro 在推荐列表中", 'plot_miyashiro' in fn_names)
    test("Mg# 在推荐列表中", 'plot_mgno' in fn_names)

    # 模拟删除 FeO 和 TFe2O3
    saved = gd._elem_data.copy()
    for k in ['FeO', 'TFe2O3']:
        if k in gd._elem_data:
            del gd._elem_data[k]

    diags2, _ = recommended_diagrams(gd, rock_type='mafic')
    fn_names2 = {fn.__name__ for fn, _ in diags2}
    test("FeO+TFe2O3 都缺时 AFM 被跳过", 'plot_afm' not in fn_names2, f"got {fn_names2}")
    test("FeO+TFe2O3 都缺时 Miyashiro 被跳过", 'plot_miyashiro' not in fn_names2)
    test("FeO+TFe2O3 都缺时 Mg# 被跳过", 'plot_mgno' not in fn_names2)

    # 恢复
    gd._elem_data = saved
    os.unlink(path)


def test_registry_self_consistency():
    """检查注册表内部自洽性：每个 spec 的 fn 有 __name__，文件名不重复。"""
    print("\n[13] 注册表内部自洽性")
    from whole_rock_core import DIAGRAM_REGISTRY

    fnames = {}
    for spec in DIAGRAM_REGISTRY:
        fn = spec.fn
        test(f"{fn.__name__} 有 __name__", hasattr(fn, '__name__'))
        if spec.filename in fnames:
            test(f"{spec.filename} 文件名不重复", False, f"冲突: {fn.__name__} vs {fnames[spec.filename]}")
        else:
            fnames[spec.filename] = fn.__name__

    mafic_count = sum(1 for d in DIAGRAM_REGISTRY if 'mafic' in d.rock_types)
    felsic_count = sum(1 for d in DIAGRAM_REGISTRY if 'felsic' in d.rock_types)
    total = len(DIAGRAM_REGISTRY)
    print(f"     注册表总计: {total} 张图 (mafic: {mafic_count}, felsic: {felsic_count})")


def test_diagram_tuples_four_fields():
    print("\n[12] 数据表结构一致性")
    for fn, desc, needed, any_of in MAFIC_DIAGRAMS + FELSIC_DIAGRAMS:
        test(f"{fn.__name__} 是四元组", isinstance(any_of, (tuple, type(None))))
        if any_of:
            test(f"  {any_of} 中无元素与 needed 重复", not set(any_of) & set(needed))


# ── 主入口 ──────────────────────────────────────────────

def main():
    global _PASS, _FAIL, _SKIP
    _PASS = _FAIL = _SKIP = 0

    print(f"🔬 whole_rock_core 验证 — SKILL_DIR={SKILL_DIR}")
    if QUICK_MODE:
        print("   ⚡ Quick 模式：跳过出图，仅做 import/registry 检查")

    # ── 环境预检 ──
    try:
        import importlib.metadata as ilm
        REQS = {'numpy': '1.24', 'matplotlib': '3.6', 'scipy': '1.10', 'openpyxl': '3.1'}
        for pkg, min_ver in REQS.items():
            ver = ilm.version(pkg)
            test(f"依赖 {pkg}>={min_ver}", list(map(int, ver.split('.'))) >= list(map(int, min_ver.split('.'))),
                 f"已安装 {ver}")
    except ImportError as e:
        test("依赖检查", False, str(e))

    pyok = sys.version_info >= (3, 10)
    test("Python >= 3.10", pyok, f"当前 {sys.version}")

    tests = [test_data_loading, test_transpose_loading,
             test_detection_limit, test_sample_filter,
             test_normalize, test_set_out_dir, test_set_style]
    if not QUICK_MODE:
        tests.append(test_plot_smoke)
    tests += [test_recommended_diagrams]
    if not QUICK_MODE:
        tests.append(test_plot_recommended)
    tests += [test_any_of_or_condition, test_diagram_tuples_four_fields,
              test_registry_self_consistency]

    for test_fn in tests:
        try:
            test_fn()
        except Exception as e:
            _FAIL += 1
            print(f"  🔴 测试异常: {test_fn.__name__}: {e}")

    print(f"\n{'='*40}")
    print(f"结果: ✅ {_PASS} / ❌ {_FAIL} / ⏭️  {_SKIP}")
    if _FAIL == 0:
        print("🎉 全部通过！")
    else:
        print(f"⚠️  有 {_FAIL} 项失败")
    return 0 if _FAIL == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
