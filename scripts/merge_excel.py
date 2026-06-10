#!/usr/bin/env python3
"""
merge_excel.py — 合并主量/微量元素 Excel 报告

功能：
- 自动检测两种常见格式：
  标准格式：元素在行、样品在列（实验室报告常见）
  转置格式：样品在行、元素在列（部分实验室报告）
- 按样品编号对齐主量和微量数据
- 自动排除标准物质行（GSR-2, 推荐值等）
- 输出标准化 4 行表头格式，供 igneous-geochemistry skill 直接读取

输出格式：
  Row 1: Sample ID | 样品编号
  Row 2:            | Major/Trace 类别（合并单元格）
  Row 3: Element    | 元素符号
  Row 4: Unit       | wt% / ppm
  Row 5+: 元素符号  | 数据

用法（直接运行）：
  python merge_excel.py <主量文件> <微量文件> [-o 输出路径]

用法（Python import）：
  from merge_excel import merge_two_files
"""

import os
import sys
import argparse

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ── 元素识别关键字 ──────────────────────────────────

MAJOR_KEYWORDS = {'SiO2', 'TiO2', 'Al2O3', 'TFe2O3', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O', 'P2O5'}
TRACE_KEYWORDS = {'Li', 'Be', 'Sc', 'Ti', 'V', 'Cr', 'Ni', 'Cu', 'Zn', 'Rb', 'Sr', 'Zr', 'Nb', 'Ba', 'La', 'Ce'}

# 输出排序顺序（优先匹配，剩余元素自动追加末尾）
MAJOR_ORDER = ['SiO2', 'TiO2', 'Al2O3', 'TFe2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O', 'P2O5', 'LOI', 'TOTAL']
TRACE_ORDER = [
    'Li', 'Be', 'Sc', 'Ti', 'V', 'Cr', 'Mn', 'Co', 'Ni', 'Cu', 'Zn', 'Ga',
    'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Sn', 'Cs', 'Ba',
    'La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu',
    'Hf', 'Ta', 'W', 'Tl', 'Pb', 'Th', 'U'
]

# ── 常用地球化学计算比值 ────────────────────────────

# 优先从包中导入，否则回退到内联定义
try:
    from igneous_wr.core.normalize import CHONDRITE
except ImportError:
    CHONDRITE = {
        'La': 0.237, 'Ce': 0.613, 'Pr': 0.0928, 'Nd': 0.457, 'Sm': 0.148,
        'Eu': 0.0563, 'Gd': 0.199, 'Tb': 0.0361, 'Dy': 0.246, 'Ho': 0.0546,
        'Er': 0.160, 'Tm': 0.0247, 'Yb': 0.161, 'Lu': 0.0247
    }

REE_ELEMENTS = ['La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu']
LREE_ELEMENTS = ['La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu']  # 轻稀土
HREE_ELEMENTS = ['Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu']  # 重稀土（兼容含 Y 的扩展）


def _get_val(data, elem, idx):
    """获取指定元素在 idx 位置的数值，返回 float 或 None"""
    vals = data.get(elem)
    if vals is None:
        return None
    v = vals[idx]
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def compute_ratios(sample_count, data):
    """
    自动计算常用地球化学比值。
    
    Args:
        sample_count: 样品数量
        data: {元素名: [数值列表]} 的字典
        
    Returns:
        [(比值名, [数值列表]), ...] 按论文附件常见顺序排列
    """
    ratios = []  # [(name, [values])]
    
    for i in range(sample_count):
        # 收集该样品所有 REE 数据
        ree_vals = {}
        all_ok = True
        for elem in REE_ELEMENTS:
            v = _get_val(data, elem, i)
            ree_vals[elem] = v
            if v is None:
                all_ok = False
        
        # ΣREE
        if all_ok and any(v is not None for v in ree_vals.values()):
            sree = sum(v for v in ree_vals.values() if v is not None)
        else:
            sree = None
        
        # LREE (La-Eu)
        lree_vals = [ree_vals[e] for e in LREE_ELEMENTS if ree_vals[e] is not None]
        lree = sum(lree_vals) if lree_vals and all(v is not None for v in lree_vals) else None
        
        # HREE (Gd-Lu)
        hree_vals = [ree_vals[e] for e in HREE_ELEMENTS if ree_vals[e] is not None]
        hree = sum(hree_vals) if hree_vals and all(v is not None for v in hree_vals) else None
        
        # LREE/HREE
        lh_ratio = (lree / hree) if (lree is not None and hree is not None and hree != 0) else None
        
        # Eu/Eu* = EuN / sqrt(SmN * GdN)
        eu_star = None
        if all_ok:
            eu_n = ree_vals['Eu'] / CHONDRITE['Eu'] if ree_vals['Eu'] else None
            sm_n = ree_vals['Sm'] / CHONDRITE['Sm'] if ree_vals['Sm'] else None
            gd_n = ree_vals['Gd'] / CHONDRITE['Gd'] if ree_vals['Gd'] else None
            if eu_n and sm_n and gd_n and sm_n > 0 and gd_n > 0:
                eu_star = round(eu_n / ((sm_n + gd_n) / 2), 3)
        
        # Ce/Ce* = CeN / sqrt(LaN * PrN)
        ce_star = None
        if all_ok:
            ce_n = ree_vals['Ce'] / CHONDRITE['Ce'] if ree_vals['Ce'] else None
            la_n = ree_vals['La'] / CHONDRITE['La'] if ree_vals['La'] else None
            pr_n = ree_vals['Pr'] / CHONDRITE['Pr'] if ree_vals['Pr'] else None
            if ce_n and la_n and pr_n and la_n > 0 and pr_n > 0:
                ce_star = round(ce_n / ((la_n + pr_n) / 2), 3)
        
        # 标准化比值
        def normalized_ratio(num_elem, denom_elem):
            nv = ree_vals[num_elem] / CHONDRITE[num_elem] if (all_ok and ree_vals[num_elem]) else None
            dv = ree_vals[denom_elem] / CHONDRITE[denom_elem] if (all_ok and ree_vals[denom_elem]) else None
            if nv is not None and dv is not None and dv != 0:
                return round(nv / dv, 3)
            return None
        
        n_la_yb = normalized_ratio('La', 'Yb')
        n_la_sm = normalized_ratio('La', 'Sm')
        n_gd_yb = normalized_ratio('Gd', 'Yb')
        n_dy_yb = normalized_ratio('Dy', 'Yb')
        n_sm_yb = normalized_ratio('Sm', 'Yb')
        n_tb_yb = normalized_ratio('Tb', 'Yb')
        n_ce_yb = normalized_ratio('Ce', 'Yb')
        n_pr_yb = normalized_ratio('Pr', 'Yb')
        n_nd_yb = normalized_ratio('Nd', 'Yb')
        n_la_nd = normalized_ratio('La', 'Nd')
        
        # 简单比值（两元素直接相除）
        def simple_ratio(a, b):
            av = _get_val(data, a, i)
            bv = _get_val(data, b, i)
            if av is not None and bv is not None and bv != 0:
                return round(av / bv, 2)
            return None
        
        zr_hf = simple_ratio('Zr', 'Hf')
        nb_ta = simple_ratio('Nb', 'Ta')
        zr_nb = simple_ratio('Zr', 'Nb')
        th_nb = simple_ratio('Th', 'Nb')
        th_la = simple_ratio('Th', 'La')
        la_nb = simple_ratio('La', 'Nb')
        la_ta = simple_ratio('La', 'Ta')
        ba_th = simple_ratio('Ba', 'Th')
        ba_nb = simple_ratio('Ba', 'Nb')
        rb_sr = simple_ratio('Rb', 'Sr')
        sr_y = simple_ratio('Sr', 'Y')
        zr_y = simple_ratio('Zr', 'Y')
        ti_y = simple_ratio('Ti', 'Y')
        ti_zr = simple_ratio('Ti', 'Zr')
        ni_co = simple_ratio('Ni', 'Co')
        cr_ni = simple_ratio('Cr', 'Ni')
        v_cr = simple_ratio('V', 'Cr')
        v_sc = simple_ratio('V', 'Sc')
        ce_pb = simple_ratio('Ce', 'Pb')
        nb_u = simple_ratio('Nb', 'U')
        sm_nd = simple_ratio('Sm', 'Nd')
        rb_ba = simple_ratio('Rb', 'Ba')
        ba_sr = simple_ratio('Ba', 'Sr')
        sr_nd = simple_ratio('Sr', 'Nd')
        
        ratios.append({
            'ΣREE': sree, 'LREE': lree, 'HREE': hree, 'LREE/HREE': lh_ratio,
            'Eu/Eu*': eu_star, 'Ce/Ce*': ce_star,
            '(La/Yb)N': n_la_yb, '(La/Sm)N': n_la_sm, '(Gd/Yb)N': n_gd_yb,
            '(Dy/Yb)N': n_dy_yb, '(Sm/Yb)N': n_sm_yb, '(Tb/Yb)N': n_tb_yb,
            '(Ce/Yb)N': n_ce_yb, '(Pr/Yb)N': n_pr_yb, '(Nd/Yb)N': n_nd_yb,
            '(La/Nd)N': n_la_nd,
            'Zr/Hf': zr_hf, 'Nb/Ta': nb_ta, 'Zr/Nb': zr_nb,
            'Th/Nb': th_nb, 'Th/La': th_la, 'La/Nb': la_nb, 'La/Ta': la_ta,
            'Ba/Th': ba_th, 'Ba/Nb': ba_nb,
            'Rb/Sr': rb_sr, 'Sr/Y': sr_y, 'Zr/Y': zr_y,
            'Ti/Y': ti_y, 'Ti/Zr': ti_zr,
            'Ni/Co': ni_co, 'Cr/Ni': cr_ni, 'V/Cr': v_cr, 'V/Sc': v_sc,
            'Ce/Pb': ce_pb, 'Nb/U': nb_u, 'Sm/Nd': sm_nd,
            'Rb/Ba': rb_ba, 'Ba/Sr': ba_sr, 'Sr/Nd': sr_nd,
        })
    
    # 转置：从 [sample_i:{name:val}] 变成 [(name, [vals])]
    ratio_names = list(ratios[0].keys()) if ratios else []
    result = []
    for name in ratio_names:
        vals = [r[name] for r in ratios]
        # 只保留至少有一个非 None 值的比值
        if any(v is not None for v in vals):
            result.append((name, vals))
    
    return result


# 需要排除的非样品行关键字
EXCLUDE_ROWS = {'样品编号', '推荐值', '以下空白', 'GSR-', '标准物质', '单位：', '单位:',
                'Sample', 'Report', 'Certified', 'Reference'}
# 标准物质前缀（开放科学：排除国际通用的岩石标准物质，用户可通过 --major-prefix 添加自己的样品前缀）
EXCLUDE_PREFIXES = ('GSR-', 'GSD-', 'GSS-', 'BHVO', 'AGV', 'BCR', 'DNC', 'W-2', 'MRG')

# 允许的样品编号前缀（默认宽松：排除标准物质后全部接受，可通过 --major-prefix CLI 参数收紧）
# 若需仅接受特定前缀，用 `python merge_excel.py major.xlsx trace.xlsx --major-prefix 23,24,KH`
SAMPLE_PREFIXES = ()
# 也允许纯字母（如 KH031-01 形式）
import re

def _is_valid_sample_id(sid):
    """判断是否为有效样品编号（排除标准物质和特殊行）"""
    if not sid or not isinstance(sid, str):
        return False
    sid = sid.strip()
    if not sid:
        return False
    # 排除关键字
    for kw in EXCLUDE_ROWS:
        if kw in sid:
            return False
    # 排除标准物质前缀
    for pref in EXCLUDE_PREFIXES:
        if sid.startswith(pref):
            return False
    # 如果用户设定了 SAMPLE_PREFIXES，只接受这些前缀
    if SAMPLE_PREFIXES:
        if any(sid.startswith(p) for p in SAMPLE_PREFIXES):
            return True
        if re.match(r'^[A-Z]{2,}\d', sid):
            return True
        return False
    # 默认宽松：排除标准物质后全部接受
    return True


# ── 工具函数 ────────────────────────────────────────

def _normalize(name):
    """归一化元素名：Unicode下标→普通数字 + 去除括号中文/空格"""
    if not name:
        return ''
    name = str(name).strip()
    # Unicode 下标 → 普通数字
    trans = str.maketrans('₂₃₁₄₅₆', '231456')
    name = name.translate(trans)
    # 去除括号及括号内中文内容（如 "Fe2O3（全铁）" → "Fe2O3"）
    name = re.sub(r'[（(][^)）]*[\u4e00-\u9fff]+[）)]', '', name)
    # 去掉残留的空括号
    name = re.sub(r'[（）()]', '', name)
    # 去掉空格和常见分隔符
    name = name.replace(' ', '').replace('_', '')
    # 统一大小写：首字母大写，其余小写（元素名标准写法）
    # 但对 SiO2、TiO2 等含数字和大小写的要保留原样
    if name and name[0].isalpha():
        name = name[0].upper() + name[1:]
    return name


# ── 格式自动检测 ────────────────────────────────────

def _detect_format(ws):
    """
    自动检测数据格式。
    返回 'standard'（元素在行，样品在列）或 'transposed'（样品在行，元素在列）。
    """
    KNOWN = MAJOR_KEYWORDS | TRACE_KEYWORDS | {'FeO', 'LOI', 'TOTAL', 'MnO', 'Co', 'Cu', 'Zn', 'Ga',
                                                 'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Sn', 'Cs', 'Ba',
                                                 'Hf', 'Ta', 'W', 'Tl', 'Pb', 'Th', 'U'}
    # 检查第1行A列：如果是"样品编号"、"Sample ID" → 转置格式
    r1c1 = str(ws.cell(1, 1).value or '').strip()
    if '样' in r1c1 or 'Sample' in r1c1 or r1c1 == '样品编号':
        return 'transposed'
    # 检查第1行的2-5列是否有已知元素名 → 标准格式
    for c in range(2, min(ws.max_column + 1, 10)):
        v = ws.cell(1, c).value
        if v and str(v).strip() in KNOWN:
            return 'standard'
    # 检查第2行起A列是否有已知元素名 → 转置格式
    for r in range(2, min(ws.max_row + 1, 10)):
        v = ws.cell(r, 1).value
        if v and str(v).strip() in KNOWN:
            return 'transposed'
    # 检查第1行A列是否看起来像样品编号（纯字母数字，非元素名） → 转置
    if r1c1 and not r1c1.startswith('KH') and re.match(r'^[A-Z0-9]', r1c1):
        return 'transposed'
    # fallback：标准格式
    return 'standard'


# ── 探索文件结构 ──────────────────────────────────

def explore_excel(path):
    """打印文件结构，返回 (wb, sheet_name)。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    print("工作表:", wb.sheetnames)

    sheet_name = None
    for s in wb.sheetnames:
        if '数据' in s:
            sheet_name = s
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[2] if len(wb.sheetnames) > 2 else wb.sheetnames[0]

    ws = wb[sheet_name]
    print(f"使用工作表: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        vals = [str(v) if v is not None else '' for v in row[:10]]
        print(f"  Row {i}: {' | '.join(vals)}")

    return wb, sheet_name


# ── 读取数据（两种格式通用） ──────────────────────

def read_worksheet(ws, sheet_label=""):
    """
    读取单个工作表，自动检测标准/转置格式。
    返回 (sample_ids, data_dict)。
    - sample_ids: [样品编号, ...]
    - data_dict: {元素名: [数值, ...], ...}，数值对应 sample_ids 顺序
    """
    fmt = _detect_format(ws)

    if fmt == 'transposed':
        return _read_transposed(ws, sheet_label)
    else:
        return _read_standard(ws, sheet_label)


def _read_transposed(ws, sheet_label=""):
    """
    转置格式：第1行样品编号+元素名，第2行起数据。
    例如：
      Row 1: 样品编号 | Li | Be | Sc | Ti | V | Cr | ...
      Row 2: 23TEQ02-1 | 29.9 | 2.1 | 17.8 | 3814 | ...
    """
    # 第1行：元素名
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(str(v).strip() if v else '')

    # 第2行起：数据，直到遇到非样品行
    samples = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        sid_str = str(sid).strip() if sid else ''
        if _is_valid_sample_id(sid_str):
            samples.append(r)
        else:
            # 如果连续3行都不是有效样品，停止
            if r > 2 and all(not _is_valid_sample_id(str(ws.cell(r+k, 1).value or '').strip())
                             for k in range(3)):
                break

    if not samples:
        raise ValueError(f"在 {sheet_label} 中找不到有效样品数据")

    sample_ids = []
    data = {h: [] for h in headers[1:] if h and h not in ('样品编号', 'Sample ID')}

    for r in samples:
        sid = str(ws.cell(r, 1).value).strip()
        sample_ids.append(sid)
        for c in range(2, len(headers) + 1):
            h = headers[c - 1]
            if h and h not in ('样品编号', 'Sample ID'):
                v = ws.cell(r, c).value
                data.setdefault(h, []).append(v)

    print(f"  [{sheet_label}] 格式: 转置, 样品数: {len(sample_ids)}, 元素数: {len([h for h in data if any(v is not None for v in data[h])])}")
    return sample_ids, data


def _read_standard(ws, sheet_label=""):
    """
    标准格式：第1行样品名横向排列，A列元素名竖向排列。
    例如：
      Row 1: 实验室编号 | 样品编号 | SiO2 | TiO2 | Al2O3 | ...
      Row 2: KH031-01 | 23TEQ02-1 | 60.56 | 0.56 | 14.66 | ...
    """
    # 第1行：列标题
    col_headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            col_headers[c] = str(v).strip()

    KNOWN = MAJOR_KEYWORDS | TRACE_KEYWORDS | {'FeO', 'LOI', 'TOTAL', 'MnO', 'Co', 'Cu', 'Zn', 'Ga',
                                                 'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Sn', 'Cs', 'Ba',
                                                 'Hf', 'Ta', 'W', 'Tl', 'Pb', 'Th', 'U'}

    # 找出元素列（已知元素名）
    elem_cols = [c for c, h in col_headers.items() if h in KNOWN]
    if not elem_cols:
        # 如果没有已知元素名，尝试用排除法：非样品编号列、非第1列
        for c in range(3, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and not any(kw in str(v) for kw in ['编', '号', 'Sample', 'ID']):
                elem_cols.append(c)

    if not elem_cols:
        raise ValueError(f"在 {sheet_label} 中找不到元素数据列")

    # 第2行起找样品数据（跳过单位行、空白行）
    samples = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 2).value  # 样品编号通常在第2列
        if not sid:
            sid = ws.cell(r, 1).value
        sid_str = str(sid).strip() if sid else ''
        # 检查这一行是否有数值数据（确认是数据行而非单位行）
        has_numeric = False
        for c in elem_cols[:3]:  # 检查前几个元素列
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                has_numeric = True
                break
        if has_numeric and _is_valid_sample_id(sid_str):
            samples.append(r)
        elif r > 5 and '以下空白' in sid_str:
            break  # 结束标记

    if not samples:
        # fallback: 如果样品编号是 KH031-01 这类数字+字母格式
        for r in range(2, min(ws.max_row + 1, 25)):
            sid = ws.cell(r, 2).value
            if sid and not isinstance(sid, (int, float)):
                sid_str = str(sid).strip()
                # 检查有数值数据
                has_num = any(isinstance(ws.cell(r, c).value, (int, float)) for c in elem_cols[:3])
                if has_num and sid_str and sid_str != '以下空白':
                    samples.append(r)

    sample_ids = []
    data = {}

    for r in samples:
        sid = str(ws.cell(r, 2).value or ws.cell(r, 1).value).strip()
        sample_ids.append(sid)

    for c in elem_cols:
        elem = col_headers[c]
        norm = _normalize(elem)
        vals = []
        for r in samples:
            v = ws.cell(r, c).value
            vals.append(v)
        data[norm] = vals

    print(f"  [{sheet_label}] 格式: 标准, 样品数: {len(sample_ids)}, 元素数: {len(data)}")
    return sample_ids, data


# ── Step 4: 合并主量和微量 ─────────────────────

def merge_major_trace(maj_data, trace_data, major_ids, trace_ids):
    """
    合并主量和微量数据，按样品编号对齐。
    以主量样品顺序为准（微量只取主量中存在的样品）。
    """
    trace_idx = {sid: i for i, sid in enumerate(trace_ids)}
    all_samples = list(major_ids)
    merged = dict(maj_data)

    for elem, trace_vals in trace_data.items():
        aligned = []
        for sid in all_samples:
            if sid in trace_idx:
                aligned.append(trace_vals[trace_idx[sid]])
            else:
                aligned.append(None)
        merged[elem] = aligned

    return all_samples, merged


# ── Step 5: 排序元素（预定义顺序 + 动态追加） ────

def sort_elements(merged_data):
    """
    按预定义顺序排序元素，未匹配的元素自动归入主量/微量。
    返回 (major_cols, trace_cols)，每个元素名用原始名称。
    """
    data_norm = {_normalize(k): k for k in merged_data}

    major_cols = []
    for e in MAJOR_ORDER:
        if e in data_norm:
            major_cols.append(data_norm.pop(e))

    trace_cols = []
    for e in TRACE_ORDER:
        if e in data_norm:
            trace_cols.append(data_norm.pop(e))

    # 剩余未匹配：含 O 的归主量，其余归微量
    for orig_name in data_norm.values():
        if 'O' in orig_name:
            major_cols.append(orig_name)
        else:
            trace_cols.append(orig_name)

    return major_cols, trace_cols


# ── Step 6: 输出标准格式 Excel ────────────────────

def write_merged_excel(all_samples, merged_data, output_path):
    """
    输出合并后的 Excel（论文附件格式：element × sample 矩阵，无 4 行表头）。

    输出格式（与论文 Table 1 一致）：
      Row 1: Sample | 样品编号横排
      Row 2+: 元素名 | 数值
    """
    major_cols, trace_cols = sort_elements(merged_data)
    all_cols = major_cols + trace_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Geochemistry'

    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_cell(row, col, value, bold=False, align='center'):
        cell = ws.cell(row, col, value)
        cell.font = Font(bold=bold, size=11)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = border

    def fmt_val(val):
        if val is None:
            return ''
        return val

    def set_precision(val):
        if not isinstance(val, (int, float)):
            return val
        if abs(val) >= 100:
            return int(round(val))
        elif abs(val) >= 10:
            return round(val, 1)
        else:
            return round(val, 2)

    # Row 1: 样品编号
    write_cell(1, 1, 'Sample', bold=True)
    for j, sid in enumerate(all_samples, 2):
        write_cell(1, j, sid)

    # Row 2+: 元素数据
    for i, elem in enumerate(all_cols):
        row = i + 2
        write_cell(row, 1, elem)
        for j, val in enumerate(merged_data[elem], 2):
            v = fmt_val(val)
            if isinstance(v, (int, float)):
                v = set_precision(v)
            write_cell(row, j, v)

    # ── 追加计算比值 ──
    next_row = len(all_cols) + 2
    computed = compute_ratios(len(all_samples), merged_data)
    ratio_count = 0
    for name, vals in computed:
        row = next_row + ratio_count
        write_cell(row, 1, name, bold=True)
        for j, v in enumerate(vals, 2):
            if v is not None:
                write_cell(row, j, round(v, 3) if isinstance(v, float) else v)
        ratio_count += 1

    print(f"  计算比值: {ratio_count} 个")

    # 列宽
    ws.column_dimensions['A'].width = 12
    from openpyxl.utils import get_column_letter
    for j in range(2, len(all_samples) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14

    wb.save(output_path)
    print(f'合并完成: {output_path}')
    print(f'  元素数: {len(all_cols)} 个')
    print(f'  样品数: {len(all_samples)}')

# ── IDE 入口 ─────────────────────────────────────

def merge_multiple_files(file_paths, output=None):
    """
    批量合并多个 Excel 文件。

    将所有文件按主量/微量自动识别后统一合并。
    支持 glob 匹配的多个文件—自动识别主量（含 major/主量 关键字或有主量元素列）
    和微量（其余文件），然后逐一合并。

    Args:
        file_paths: Excel 文件路径列表
        output: 输出路径（默认第一个文件所在目录）

    Returns:
        str — 输出文件路径
    """
    if not file_paths:
        print("[merge_excel] 没有提供文件")
        return None

    if len(file_paths) == 2:
        # 2个文件：直接调用原合并逻辑
        return merge_two_files(file_paths[0], file_paths[1], output)

    print(f"[merge_excel] 批量合并 {len(file_paths)} 个文件...")
    for p in file_paths:
        print(f"  - {os.path.basename(p)}")

    # 自动分类：按文件名关键字和列内容判断主量/微量
    major_files = []
    trace_files = []
    unclassified = []

    for fp in file_paths:
        base = os.path.basename(fp).lower()
        if 'major' in base or '主量' in base or '常量' in base:
            major_files.append(fp)
        elif 'trace' in base or '微量' in base or '稀土' in base or 'ree' in base:
            trace_files.append(fp)
        elif 'stack' in base or '稀土' in base or '不分' in base:
            # stack/稀土很可能包含全部数据，单独处理
            unclassified.append(fp)
        elif '合并' in base or 'merged' in base or 'geochem' in base:
            print(f"  跳过（可能是之前合并的输出）: {os.path.basename(fp)}")
            continue
        else:
            unclassified.append(fp)

    # 尝试从未分类的文件中通过列名识别
    KNOWN_MAJOR = {'SiO2', 'TiO2', 'Al2O3', 'TFe2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O', 'P2O5'}
    for fp in list(unclassified):
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb[wb.sheetnames[0]]
            row1_vals = [str(ws.cell(1, c).value or '') for c in range(1, min(ws.max_column+1, 20))]
            row2_vals = [str(ws.cell(2, c).value or '') for c in range(1, min(ws.max_column+1, 20))]
            all_vals = set(row1_vals + row2_vals)
            major_in = len(KNOWN_MAJOR & all_vals)
            # 如果有 3 个以上主量元素关键词 → 主量文件
            if major_in >= 3:
                major_files.append(fp)
                unclassified.remove(fp)
                print(f"  [自动识别] {os.path.basename(fp)} → 主量文件")
            else:
                # 检查是否有微量元素列
                trace_in = sum(1 for v in all_vals if v in TRACE_KEYWORDS)
                if trace_in >= 3:
                    trace_files.append(fp)
                    unclassified.remove(fp)
                    print(f"  [自动识别] {os.path.basename(fp)} → 微量文件")
        except Exception:
            pass

    # 剩余未分类的作为微量处理
    trace_files.extend(unclassified)

    if not major_files and not trace_files:
        print("[merge_excel] 无法自动分类文件，按传入顺序第1个为主量，其余为微量")
        major_files = [file_paths[0]]
        trace_files = file_paths[1:]

    print(f"\n  主量文件: {len(major_files)} 个")
    for p in major_files:
        print(f"    - {os.path.basename(p)}")
    print(f"  微量文件: {len(trace_files)} 个")
    for p in trace_files:
        print(f"    - {os.path.basename(p)}")

    # 处理多个主量文件时，合并主量（取并集）
    if len(major_files) > 1:
        print("\n  合并多个主量文件...")
        combined_major = {}
        all_maj_ids = []
        for i, fp in enumerate(major_files):
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb[wb.sheetnames[0]]
            ids, data = read_worksheet(ws, f"主量{i+1}")
            if i == 0:
                all_maj_ids = ids
                combined_major = data
            else:
                for elem, vals in data.items():
                    if elem not in combined_major:
                        # 对新元素尝试对齐样品
                        combined_major[elem] = _align_by_sample(all_maj_ids, ids, vals)
                    else:
                        # 已有元素以第一个文件为准
                        pass
        maj_data = combined_major
        maj_ids = all_maj_ids
        print(f"    主量合并后: {len(maj_ids)} 样品, {len(maj_data)} 元素")
    else:
        wb_maj = openpyxl.load_workbook(major_files[0], data_only=True)
        ws_maj = wb_maj[wb_maj.sheetnames[0]]
        maj_ids, maj_data = read_worksheet(ws_maj, "主量")

    # 合并所有微量文件
    if len(trace_files) > 1:
        print("\n  合并多个微量文件...")
        combined_trace = {}
        all_tr_ids = []
        for i, fp in enumerate(trace_files):
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb[wb.sheetnames[0]]
            ids, data = read_worksheet(ws, f"微量{i+1}")
            if i == 0:
                all_tr_ids = ids
                combined_trace = data
            else:
                for elem, vals in data.items():
                    if elem not in combined_trace:
                        combined_trace[elem] = _align_by_sample(all_tr_ids, ids, vals)
        trace_data = combined_trace
        trace_ids = all_tr_ids
        print(f"    微量合并后: {len(trace_ids)} 样品, {len(trace_data)} 元素")
    else:
        wb_trace = openpyxl.load_workbook(trace_files[0], data_only=True)
        ws_trace = wb_trace[wb_trace.sheetnames[0]]
        trace_ids, trace_data = read_worksheet(ws_trace, "微量")

    # 最终合并
    print("\n  合并主量+微量...")
    all_samples, merged = merge_major_trace(maj_data, trace_data, maj_ids, trace_ids)

    output_path = output
    if output_path is None:
        output_path = os.path.join(os.path.dirname(os.path.abspath(file_paths[0])),
                                    'merged_geochemistry.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    write_merged_excel(all_samples, merged, output_path)
    return output_path


def _align_by_sample(target_ids, source_ids, source_vals):
    """将 source 的数据按 target_ids 顺序对齐。"""
    id_to_val = {sid: v for sid, v in zip(source_ids, source_vals)}
    result = []
    for sid in target_ids:
        result.append(id_to_val.get(sid, None))
    return result


def _merge_pair(major_path, trace_path, output=None):
    """合并一对主量/微量文件，返回输出路径。"""
    if output is None:
        output = os.path.join(os.path.dirname(os.path.abspath(major_path)),
                               'merged_geochemistry.xlsx')
    os.makedirs(os.path.dirname(output), exist_ok=True)

    print("=" * 50)
    print("主量元素文件")
    print("=" * 50)
    wb_maj = openpyxl.load_workbook(major_path, data_only=True)
    sn_maj = None
    for s in wb_maj.sheetnames:
        if '数据' in s:
            sn_maj = s
            break
    if sn_maj is None:
        sn_maj = wb_maj.sheetnames[2] if len(wb_maj.sheetnames) > 2 else wb_maj.sheetnames[0]
    ws_maj = wb_maj[sn_maj]
    print(f"工作表: {ws_maj.title}")
    maj_ids, maj_data = read_worksheet(ws_maj, sheet_label="主量")
    print(f"主量样品: {maj_ids}")
    print(f"主量元素: {list(maj_data.keys())}")

    print()
    print("=" * 50)
    print("微量元素文件")
    print("=" * 50)
    wb_trace = openpyxl.load_workbook(trace_path, data_only=True)
    sn_trace = None
    for s in wb_trace.sheetnames:
        if '数据' in s:
            sn_trace = s
            break
    if sn_trace is None:
        sn_trace = wb_trace.sheetnames[2] if len(wb_trace.sheetnames) > 2 else wb_trace.sheetnames[0]
    ws_trace = wb_trace[sn_trace]
    print(f"工作表: {ws_trace.title}")
    trace_ids, trace_data = read_worksheet(ws_trace, sheet_label="微量")
    print(f"微量样品: {trace_ids}")
    print(f"微量元素: {list(trace_data.keys())}")

    print()
    print("=" * 50)
    print("合并中...")
    all_samples, merged = merge_major_trace(maj_data, trace_data, maj_ids, trace_ids)
    write_merged_excel(all_samples, merged, output)
    return output


def merge_two_files(major_path, trace_path, output=None):
    """合并两个文件（保持向后兼容）。"""
    return _merge_pair(major_path, trace_path, output)


# ── CLI 入口 ─────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='合并主量/微量元素 Excel 报告')
    parser.add_argument('major', help='主量元素 Excel 文件路径')
    parser.add_argument('trace', help='微量元素 Excel 文件路径')
    parser.add_argument('-o', '--output', default=None,
                        help='输出路径（默认: 主量文件同目录下 merged_geochemistry.xlsx）')
    parser.add_argument('--sheet-major', default=None,
                        help='主量文件的工作表名（跳过自动选择）')
    parser.add_argument('--sheet-trace', default=None,
                        help='微量文件的工作表名（跳过自动选择）')
    parser.add_argument('--major-prefix', default=None,
                        help='主量文件的样品编号前缀（强制切换格式识别）')
    parser.add_argument('--trace-prefix', default=None,
                        help='微量文件的样品编号前缀（强制切换格式识别）')
    parser.add_argument('--glob', nargs='+', default=None,
                        help='批量合并模式：传入多个文件路径或 glob 模式')
    args = parser.parse_args()

    # 批量合并模式
    if args.glob:
        import glob as _glob
        file_paths = []
        for pat in args.glob:
            file_paths.extend(_glob.glob(pat))
        if not file_paths:
            print("[merge_excel] 未匹配到任何文件")
            return
        merge_multiple_files(file_paths, output=args.output)
        return

    # 处理前缀
    if args.major_prefix:
        global SAMPLE_PREFIXES
        SAMPLE_PREFIXES = SAMPLE_PREFIXES + tuple(args.major_prefix.split(','))
    if args.trace_prefix:
        # 已在 _merge_pair 中处理
        pass

    merge_two_files(args.major, args.trace, output=args.output)


if __name__ == '__main__':
    main()
