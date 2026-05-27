"""
core/data.py — GeochemData 数据容器与推荐图件调度

注意：内部 import 倒挂层（_style 等是 scripts/ 目录下的普通模块，非包内模块）
"""
import numpy as np
import openpyxl

from igneous_geochem.io.excel import (
    find_excel, parse_value, ELEM_ALIAS, KNOWN_ELEMENTS,
    DL_STRATEGY_HALF, DL_STRATEGY_ZERO, DL_STRATEGY_NAN,
)


class GeochemData:
    """
    地球化学数据容器。
    读取 merged Excel，自动检测格式，提供元素值查询。

    用法：
        gd = GeochemData('path/to/data.xlsx', sample_filter='24TJ02')
        sio2 = gd.get('SiO2')       # np.array，已筛选
        labels = gd.labels           # 样品名列表
        all_labels = gd.all_labels   # 全部样品名
    """

    def __init__(self, path=None, sample_filter=None, dl_strategy='half', sheet=None,
                 strict_filter=True):
        """
        Args:
            path: Excel 路径，None 则自动发现
            sample_filter: 样品筛选字符串（包含即选中），None 则全选
            dl_strategy: 检测限解析策略 'half'|'zero'|'nan'
            sheet: 工作表名，None 则自动搜索
            strict_filter: True 则空匹配报错，False 则回退到全部样品
        """
        self.path = find_excel(path)
        self.sample_filter = sample_filter
        self.dl_strategy = dl_strategy
        self._load(sheet)
        # 保存全量数据副本并设置当前 view
        self._all_elem_data = {k: v.copy() for k, v in self._elem_data.items()}
        self.idxs = list(range(len(self.all_labels)))
        self.labels = list(self.all_labels)
        self._init_groups()
        if sample_filter:
            self._filter(sample_filter, strict=strict_filter)
        self._report_missing()

    def _load(self, sheet):
        wb = openpyxl.load_workbook(self.path, data_only=True)

        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            for name in ['Geochemistry', 'geochemistry', 'Sheet1']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            else:
                ws = wb[wb.sheetnames[0]]
                print(f"[whole_rock] 使用工作表: {ws.title}")

        r1_c2 = ws.cell(1, 2).value
        r2_c1 = ws.cell(2, 1).value

        transposed = False
        if r1_c2 is not None and str(r1_c2).strip() in KNOWN_ELEMENTS:
            transposed = True
        elif r2_c1 is not None and str(r2_c1).strip() in KNOWN_ELEMENTS:
            transposed = False
        else:
            transposed = any(
                str(ws.cell(1, c).value or '').strip() in KNOWN_ELEMENTS
                for c in range(2, min(ws.max_column + 1, 20))
            )

        if transposed:
            self._load_transposed(ws)
        else:
            self._load_standard(ws)

    def _load_standard(self, ws):
        """标准格式：Row 1=样品名，Col A=元素名"""
        from _normalize import REE_ORDER
        KNOWN_ELEMENTS_SET = set(KNOWN_ELEMENTS)

        sample_cols = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip() != '':
                sample_cols.append(c)
            else:
                if all(ws.cell(1, c + k).value is None for k in range(1, 3)):
                    break

        self.all_labels = [ws.cell(1, c).value for c in sample_cols]

        data_start = None
        litho_row = None
        for r in range(2, min(ws.max_row + 1, 8)):
            v = ws.cell(r, 1).value
            if v is not None:
                s = str(v).strip()
                if s in KNOWN_ELEMENTS_SET:
                    data_start = r
                    break
                elif s == 'Lithology':
                    litho_row = r

        if data_start is None:
            data_start = 4

        self._lithology = None
        if litho_row is not None:
            self._lithology = [ws.cell(litho_row, c).value for c in sample_cols]

        elem_row = {}
        for r in range(data_start, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None:
                elem = str(v).strip()
                if elem and elem not in ('', 'Element', 'Unit', 'Major elements', 'Trace elements'):
                    if elem in ('ΣREE', '∑REE', 'Eu/Eu*', '(La/Yb)N', '(Dy/Yb)N', '(La/Sm)N', '(Gd/Yb)N'):
                        continue
                    elem_row[elem] = r

        self._elem_data = {}
        for elem, r in elem_row.items():
            canon = ELEM_ALIAS.get(elem, elem)
            if canon not in self._elem_data:
                self._elem_data[canon] = np.array(
                    [parse_value(ws.cell(r, c).value, self.dl_strategy) for c in sample_cols]
                )

        print(f"[whole_rock] 标准格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素"
              f"{' (含 Lithology)' if self._lithology else ''}")

    def _load_transposed(self, ws):
        """转置格式：Row 1 横铺元素名，Row 4+ 每行一个样品"""
        KNOWN_SET = set(KNOWN_ELEMENTS)

        elem_count = 0
        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()
            if v in KNOWN_SET:
                elem_count += 1
            if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                break

        if elem_count < 3:
            print(f"[whole_rock] ⚠️ 转置检测：Row 1 仅识别 {elem_count} 个元素列，按标准格式处理")
            self._load_standard(ws)
            return

        elem_cols = {}
        sample_labels = []
        in_elem_section = True

        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()

            if in_elem_section:
                if v in KNOWN_SET:
                    elem_cols[v] = c
                    continue
                elif v and v not in ('', 'Major elements (wt%)', 'Trace elements (ppm)',
                                     'Minor elements (ppm)', 'Element'):
                    in_elem_section = False
                    sample_labels.append(c)
                    continue
                elif not v:
                    if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                        break
                    continue
            else:
                if v:
                    sample_labels.append(c)
                elif all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                    break

        for r in range(4, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue
            if label.upper().startswith(('BCR', 'BHVO', 'AGV', 'GSP', 'G-2', 'JB-', 'JA-',
                                          'JG-', 'JP-', 'JR-', 'GSB', 'NIST', 'SRM', 'SY-',
                                          'MRG', 'PCC', 'DTS', 'W-2', 'DNC', 'BIR', 'UB-N',
                                          'ACE', 'SARM', 'IMA', 'SDC', 'DL-', 'GSR', 'GSS')):
                continue
            sample_labels.append(r)

        self.all_labels = []
        self._elem_data = {}

        for r in range(4, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue
            if label.upper().startswith(('BCR', 'BHVO', 'AGV', 'GSP', 'G-2', 'JB-', 'JA-',
                                          'JG-', 'JP-', 'JR-', 'GSB', 'NIST', 'SRM', 'SY-',
                                          'MRG', 'PCC', 'DTS', 'W-2', 'DNC', 'BIR', 'UB-N',
                                          'ACE', 'SARM', 'IMA', 'SDC', 'DL-', 'GSR', 'GSS')):
                continue

            self.all_labels.append(label)
            for elem, c in elem_cols.items():
                canon = ELEM_ALIAS.get(elem, elem)
                val = parse_value(ws.cell(r, c).value, self.dl_strategy)
                if canon not in self._elem_data:
                    self._elem_data[canon] = []
                self._elem_data[canon].append(val)

        for k in self._elem_data:
            self._elem_data[k] = np.array(self._elem_data[k])

        print(f"[whole_rock] 转置格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素")

    def _filter(self, keyword, strict=True):
        idxs = [i for i, l in enumerate(self.all_labels) if keyword in str(l)]
        if not idxs:
            if strict:
                raise ValueError(
                    f"筛选关键字 '{keyword}' 未匹配任何样品 "
                    f"(共 {len(self.all_labels)} 个: {self.all_labels[:5]}...)。"
                    f"请检查关键字，或传 sample_filter=None 全选。")
            else:
                print(f"[whole_rock] ⚠️ 筛选关键字 '{keyword}' 未匹配任何样品，回退到全部样品")
                self.idxs = list(range(len(self.all_labels)))
                self.labels = list(self.all_labels)
        else:
            self.idxs = idxs
            self.labels = [self.all_labels[i] for i in idxs]
            print(f"[whole_rock] 筛选 '{keyword}'：{len(idxs)}/{len(self.all_labels)} 样品")

        self._elem_data = {}
        for elem in self._all_elem_data:
            full_arr = self._all_elem_data[elem]
            self._elem_data[elem] = np.array([full_arr[i] for i in self.idxs])
        self.groups = self.infer_groups()

    def infer_groups(self):
        import re
        groups = []
        for lbl in self.labels:
            lbl_str = str(lbl).strip()
            m = re.match(r'^([A-Za-z0-9]+)-', lbl_str)
            if m:
                groups.append(m.group(1))
            else:
                groups.append(lbl_str)
        return groups

    def _init_groups(self):
        self.groups = self.infer_groups()

    def _report_missing(self):
        pass

    def get(self, elem_name):
        canon = ELEM_ALIAS.get(elem_name, elem_name)
        if canon in self._elem_data:
            return self._elem_data[canon]
        if elem_name in self._elem_data:
            return self._elem_data[elem_name]
        print(f"[whole_rock] ⚠️ 元素 '{elem_name}' 未找到")
        return np.full(len(self.labels), np.nan)

    def check_elements(self, *elems, strict=False):
        missing = []
        for e in elems:
            canon = ELEM_ALIAS.get(e, e)
            if canon not in self._elem_data and e not in self._elem_data:
                missing.append(e)
        if missing:
            if strict:
                print(f"[whole_rock] ❌ 缺失关键元素: {missing}，无法出图")
            else:
                print(f"[whole_rock] ⚠️ 缺失元素: {missing}，图中对应位置将为空")
        return missing
