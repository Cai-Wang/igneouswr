"""
core/data.py — GeochemData 数据容器与推荐图件调度

注意：内部 import 倒挂层（_style 等是 scripts/ 目录下的普通模块，非包内模块）
"""
import numpy as np
import openpyxl

from igneous_wr.io.excel import (
    find_excel, parse_value, ELEM_ALIAS, KNOWN_ELEMENTS,
    DL_STRATEGY_HALF, DL_STRATEGY_ZERO, DL_STRATEGY_NAN,
)

# 自动排除的标准物质/参考样前缀
_REFERENCE_PREFIXES = ('BCR', 'BHVO', 'AGV', 'GSP', 'G-2', 'JB-', 'JA-',
                       'JG-', 'JP-', 'JR-', 'GSB', 'NIST', 'SRM', 'SY-',
                       'MRG', 'PCC', 'DTS', 'W-2', 'DNC', 'BIR', 'UB-N',
                       'ACE', 'SARM', 'IMA', 'SDC', 'DL-', 'GSR', 'GSS')


class GeochemData:
    """
    地球化学数据容器。
    读取 merged Excel，自动检测格式，提供元素值查询。

    用法：
        gd = GeochemData('path/to/data.xlsx', sample_filter='24TJ02')
        sio2 = gd.get('SiO2')       # np.array，已筛选
        labels = gd.labels           # 样品名列表
        all_labels = gd.all_labels   # 全部样品名
    """

    def __init__(self, path=None, sample_filter=None, dl_strategy='half', sheet=None,
                 strict_filter=True):
        """
        Args:
            path: Excel 路径，None 则自动发现
            sample_filter: 样品筛选字符串（包含即选中），None 则全选
            dl_strategy: 检测限解析策略 'half'|'zero'|'nan'
            sheet: 工作表名，None 则自动搜索
            strict_filter: True 则空匹配报错，False 则回退到全部样品
        """
        self.path = find_excel(path)
        self.sample_filter = sample_filter
        self.dl_strategy = dl_strategy
        self._load(sheet)
        # 保存全量数据副本并设置当前 view
        self._all_elem_data = {k: v.copy() for k, v in self._elem_data.items()}
        self.idxs = list(range(len(self.all_labels)))
        self.labels = list(self.all_labels)
        self._init_groups()
        if sample_filter:
            self._filter(sample_filter, strict=strict_filter)
        self._report_missing()
        self._post_load_sanity_check()

    def _detect_layout(self, ws):
        """
        检测 Excel 数据布局，返回 (mode, reason)。

        mode='standard' → Row1=样品名，Col A=元素名
        mode='transposed' → Row1=元素名，Col A=样品名
        mode='wide' → Row1=元素名横铺(A1除外)，Col A=样品名(即你的标准宽表)

        通过统计 Row1 和 Col A 两轴的 KNOWN_ELEMENTS 匹配数来判断，
        而不是只靠单个格子。
        """
        KNOWN_SET = set(KNOWN_ELEMENTS)

        # ── 统计 Row1（第1行第2列起）的元素列数 ──
        row1_elem_count = 0
        row1_sample_hint = False
        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()
            if v in KNOWN_SET:
                row1_elem_count += 1
            elif v.lower() in ('sample', 'sample id', 'sample_id', 'id', 'name'):
                row1_sample_hint = True

        # ── 统计 Col A（第2行起）的元素名数量 ──
        col_a_elem_count = 0
        col_a_non_blank = 0
        for r in range(2, min(ws.max_row + 1, 20)):
            v = str(ws.cell(r, 1).value or '').strip()
            if v:
                col_a_non_blank += 1
                if v in KNOWN_SET:
                    col_a_elem_count += 1

        # ── 判断 ──
        # 情况 A: Row1 全是元素名（比如 B1=Al2O3, C1=Ba...），A1 通常是 "Sample" 或无意义
        #         → 这是最常见的"宽表"格式：Row1=元素横铺，Col A=样品名
        #         判断标准：Row1 元素列多，Col A 几乎没有元素名
        if row1_elem_count >= 3 and col_a_elem_count == 0:
            return 'wide', f"Row1 含 {row1_elem_count} 个元素，Col A 无元素"

        # 情况 B: Row1 有样品名（A1=无关, B1=24TJ02-1...），Col A 有元素名
        #         → 这是"标准转置"格式：Row1=样品横铺，Col A=元素名
        if col_a_elem_count >= 3 and row1_elem_count == 0:
            return 'standard', f"Col A 含 {col_a_elem_count} 个元素，Row1 无元素"

        # 情况 C: 两者都有——Row1 有些元素也有些样品，或者都少
        #         看 Row1 中元素占比 vs Col A 中元素占比
        row1_total = min(ws.max_column - 1, 80)  # 检查的有效列
        row1_ratio = row1_elem_count / max(row1_total, 1)
        col_a_ratio = col_a_elem_count / max(col_a_non_blank, 1)

        if col_a_ratio > row1_ratio and col_a_elem_count >= row1_elem_count:
            return 'standard', f"Col A 元素占比 {col_a_ratio:.0%} > Row1 {row1_ratio:.0%}"
        elif row1_ratio >= col_a_ratio and row1_elem_count >= 3:
            return 'wide', f"Row1 元素占比 {row1_ratio:.0%} >= Col A {col_a_ratio:.0%}"
        elif row1_sample_hint and row1_elem_count >= 3:
            return 'wide', f"A1 有 Sample 标记 + Row1 含 {row1_elem_count} 个元素"

        # 情况 D: 都很少，fallback 到老逻辑——看 B1
        if row1_elem_count >= 1:
            return 'transposed', f"fallback: Row1 含 {row1_elem_count} 个元素"
        return 'standard', f"fallback: 默认标准格式"

    def _load(self, sheet):
        wb = openpyxl.load_workbook(self.path, data_only=True)
        self._raw_max_row = None  # 留给事后校验用

        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            for name in ['Geochemistry', 'geochemistry', 'Sheet1']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            else:
                ws = wb[wb.sheetnames[0]]
                print(f"[IgneousWR] 使用工作表: {ws.title}")

        self._raw_max_row = ws.max_row
        self._raw_max_col = ws.max_column

        mode, reason = self._detect_layout(ws)
        self._detected_mode = mode
        print(f"[IgneousWR] 布局检测: {mode} ({reason})")

        if mode == 'wide':
            self._load_wide(ws)
        elif mode == 'transposed':
            self._load_transposed(ws)
        else:
            self._load_standard(ws)

    def _load_standard(self, ws):
        """标准格式：Row 1=样品名，Col A=元素名"""
        from igneous_wr.core.normalize import REE_ORDER
        KNOWN_ELEMENTS_SET = set(KNOWN_ELEMENTS)

        sample_cols = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip() != '':
                sample_cols.append(c)
            else:
                if all(ws.cell(1, c + k).value is None for k in range(1, 3)):
                    break

        self.all_labels = [ws.cell(1, c).value for c in sample_cols]

        data_start = None
        litho_row = None
        for r in range(2, min(ws.max_row + 1, 8)):
            v = ws.cell(r, 1).value
            if v is not None:
                s = str(v).strip()
                if s in KNOWN_ELEMENTS_SET:
                    data_start = r
                    break
                elif s == 'Lithology':
                    litho_row = r

        if data_start is None:
            data_start = 4

        self._lithology = None
        if litho_row is not None:
            self._lithology = [ws.cell(litho_row, c).value for c in sample_cols]

        elem_row = {}
        for r in range(data_start, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None:
                elem = str(v).strip()
                if elem and elem not in ('', 'Element', 'Unit', 'Major elements', 'Trace elements'):
                    if elem in ('ΣREE', '∑REE', 'Eu/Eu*', '(La/Yb)N', '(Dy/Yb)N', '(La/Sm)N', '(Gd/Yb)N'):
                        continue
                    elem_row[elem] = r

        self._elem_data = {}
        for elem, r in elem_row.items():
            canon = ELEM_ALIAS.get(elem, elem)
            if canon not in self._elem_data:
                self._elem_data[canon] = np.array(
                    [parse_value(ws.cell(r, c).value, self.dl_strategy) for c in sample_cols]
                )

        print(f"[IgneousWR] 标准格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素"
              f"{' (含 Lithology)' if self._lithology else ''}")

    def _load_wide(self, ws):
        """宽表格式：Row 1=元素名横铺(A1=Sample或类似标记)，Col A=样品名

        这是最常见的地球化学Excel表格布局：
          A1=Sample   B1=Al2O3   C1=Ba   D1=CaO ...
          A2=24TJ02-1 B2=8.119   C2=56.2 ...
          A3=24TJ02-2 ...
        """
        KNOWN_SET = set(KNOWN_ELEMENTS)

        # 找元素列——Row1 中匹配 KNOWN_ELEMENTS 的列
        elem_cols = {}
        for c in range(2, ws.max_column + 1):
            v = str(ws.cell(1, c).value or '').strip()
            if v in KNOWN_SET:
                elem_cols[v] = c

        if not elem_cols:
            print("[IgneousWR] ⚠️ 宽表格式未在 Row1 找到元素列，回退到 transposed")
            self._load_transposed(ws)
            return

        # 读样品——Col A 从第2行起
        sample_names = []
        elem_data = {}
        for r in range(2, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label:
                continue
            # 跳过标准样和 RE 参考位置的标记
            if label.upper().startswith(_REFERENCE_PREFIXES):
                continue
            sample_names.append(label)
            for elem, col in elem_cols.items():
                canon = ELEM_ALIAS.get(elem, elem)
                val = parse_value(ws.cell(r, col).value, self.dl_strategy)
                if canon not in elem_data:
                    elem_data[canon] = []
                elem_data[canon].append(val)

        self.all_labels = sample_names
        self._elem_data = {}
        for k in elem_data:
            self._elem_data[k] = np.array(elem_data[k])
        self._lithology = None

        print(f"[IgneousWR] 宽表格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素")

    def _post_load_sanity_check(self):
        """加载完成后的一致性检查"""
        issues = []

        # 检查样品数是否合理（和文件的物理行数对比）
        # 只在 wide 模式下做：因为 wide 每行 = 一个样品
        if self._raw_max_row and self._detected_mode == 'wide':
            file_data_rows = self._raw_max_row - 1  # 去掉表头
            loaded = len(self.all_labels)
            if loaded < file_data_rows * 0.6:
                issues.append(
                    f"样品数异常：文件有 ~{file_data_rows} 行数据，只读到了 {loaded} 个样品。"
                    f"数据读取可能不完整。"
                )

        # 检查缺失的关键元素
        core_elements = ['SiO2', 'Al2O3', 'CaO', 'MgO', 'Na2O', 'K2O', 'FeO', 'TiO2']
        missing_core = [e for e in core_elements if e not in self._elem_data
                        and e not in self._elem_data.get(ELEM_ALIAS.get(e, e), {})]
        # 实际上用 get 检查
        missing_actual = []
        for e in core_elements:
            canon = ELEM_ALIAS.get(e, e)
            # 检查任何别名
            found = False
            for key in self._elem_data:
                if key == e or key == canon:
                    found = True
                    break
                if ELEM_ALIAS.get(key, key) == canon:
                    found = True
                    break
            if not found:
                missing_actual.append(e)
        if len(missing_actual) >= 4:
            issues.append(
                f"主量元素大面积缺失 ({len(missing_actual)}/8): {', '.join(missing_actual[:6])}..."
                f" 数据读取可能有误。"
            )

        for issue in issues:
            print(f"[IgneousWR] ⚠️ {issue}")
        return issues

    def _load_transposed(self, ws):
        """转置格式：Row 1 横铺元素名，Row 2+ 每行一个样品"""
        KNOWN_SET = set(KNOWN_ELEMENTS)

        elem_count = 0
        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()
            if v in KNOWN_SET:
                elem_count += 1
            if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                break

        if elem_count < 3:
            print(f"[IgneousWR] ⚠️ 转置检测：Row 1 仅识别 {elem_count} 个元素列，按标准格式处理")
            self._load_standard(ws)
            return

        elem_cols = {}
        sample_labels = []
        in_elem_section = True

        for c in range(2, min(ws.max_column + 1, 81)):
            v = str(ws.cell(1, c).value or '').strip()

            if in_elem_section:
                if v in KNOWN_SET:
                    elem_cols[v] = c
                    continue
                elif v and v not in ('', 'Major elements (wt%)', 'Trace elements (ppm)',
                                     'Minor elements (ppm)', 'Element'):
                    in_elem_section = False
                    sample_labels.append(c)
                    continue
                elif not v:
                    if all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                        break
                    continue
            else:
                if v:
                    sample_labels.append(c)
                elif all(ws.cell(1, c + k).value is None for k in range(1, 4)):
                    break

        # 自适应数据起始行：从 Row 2 开始，跳过空行和标记行
        data_start = 2
        for r in range(2, min(ws.max_row + 1, 10)):
            v = str(ws.cell(r, 1).value or '').strip()
            if not v or v.lower() in ('element', 'unit', 'sample', 'major elements (wt%)',
                                       'trace elements (ppm)', 'minor elements (ppm)'):
                data_start = r + 1
            else:
                break

        for r in range(data_start, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue
            if label.upper().startswith(_REFERENCE_PREFIXES):
                continue
            sample_labels.append(r)

        self.all_labels = []
        self._elem_data = {}

        for r in range(data_start, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or '').strip()
            if not label or label.lower() in ('', 'element', 'unit', 'sample',
                                               'major elements (wt%)', 'trace elements (ppm)',
                                               'minor elements (ppm)'):
                continue
            if label.upper().startswith(_REFERENCE_PREFIXES):
                continue

            self.all_labels.append(label)
            for elem, c in elem_cols.items():
                canon = ELEM_ALIAS.get(elem, elem)
                val = parse_value(ws.cell(r, c).value, self.dl_strategy)
                if canon not in self._elem_data:
                    self._elem_data[canon] = []
                self._elem_data[canon].append(val)

        for k in self._elem_data:
            self._elem_data[k] = np.array(self._elem_data[k])

        print(f"[IgneousWR] 转置格式，{len(self.all_labels)} 样品，{len(self._elem_data)} 元素")

    def _filter(self, keyword, strict=True):
        idxs = [i for i, l in enumerate(self.all_labels) if keyword in str(l)]
        if not idxs:
            if strict:
                raise ValueError(
                    f"筛选关键字 '{keyword}' 未匹配任何样品 "
                    f"(共 {len(self.all_labels)} 个: {self.all_labels[:5]}...)。"
                    f"请检查关键字，或传 sample_filter=None 全选。")
            else:
                print(f"[IgneousWR] ⚠️ 筛选关键字 '{keyword}' 未匹配任何样品，回退到全部样品")
                self.idxs = list(range(len(self.all_labels)))
                self.labels = list(self.all_labels)
        else:
            self.idxs = idxs
            self.labels = [self.all_labels[i] for i in idxs]
            print(f"[IgneousWR] 筛选 '{keyword}'：{len(idxs)}/{len(self.all_labels)} 样品")

        self._elem_data = {}
        for elem in self._all_elem_data:
            full_arr = self._all_elem_data[elem]
            self._elem_data[elem] = np.array([full_arr[i] for i in self.idxs])
        self.groups = self.infer_groups()

    def infer_groups(self):
        import re
        groups = []
        for lbl in self.labels:
            lbl_str = str(lbl).strip()

            # 1) 连字符前缀: MT-01-A1 → MT
            m = re.match(r'^([A-Za-z0-9]+)-', lbl_str)
            if m:
                groups.append(m.group(1))
                continue

            # 2) 字母+数字+字母核心，去掉末尾点号数字: MT01A1 → MT01A
            m = re.match(r'^([A-Za-z]+[0-9]+[A-Za-z]+)\d*$', lbl_str)
            if m:
                groups.append(m.group(1))
                continue

            # 3) 回退: 整个样品名作组名
            groups.append(lbl_str)
        return groups

    def _init_groups(self):
        self.groups = self.infer_groups()

    def _report_missing(self):
        pass

    def get(self, elem_name):
        canon = ELEM_ALIAS.get(elem_name, elem_name)
        if canon in self._elem_data:
            return self._elem_data[canon]
        if elem_name in self._elem_data:
            return self._elem_data[elem_name]
        # ── Ti 自动回退：从 TiO₂ (wt%) 换算为 Ti (ppm) ──
        if elem_name == 'Ti' and 'TiO2' in self._elem_data:
            from igneous_wr.core.chem import tio2_to_ti_ppm
            return tio2_to_ti_ppm(self._elem_data['TiO2'])
        print(f"[IgneousWR] ⚠️ 元素 '{elem_name}' 未找到")
        return np.full(len(self.labels), np.nan)

    def check_elements(self, *elems, strict=False):
        missing = []
        for e in elems:
            canon = ELEM_ALIAS.get(e, e)
            if canon in self._elem_data or e in self._elem_data:
                continue
            # Ti→TiO₂ 回退
            if e == 'Ti' and 'TiO2' in self._elem_data:
                continue
            missing.append(e)
        if missing:
            if strict:
                print(f"[IgneousWR] ❌ 缺失关键元素: {missing}，无法出图")
            else:
                print(f"[IgneousWR] ⚠️ 缺失元素: {missing}，图中对应位置将为空")
        return missing
